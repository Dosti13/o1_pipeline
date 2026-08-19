"""
Material Matcher — Exact + Fully Dynamic Fuzzy Matching

NO HARDCODED COLUMN NAMES ANYWHERE.
Every column — yours and the customer's — is found automatically
by scanning column headers with pattern lists.

HOW FUZZY WORKS (same logic for ALL files including QNIE):
  Step 1 — Find CATEGORY columns in both files  (patterns: category, group, family, mgrp, dept …)
  Step 2 — Find DESCRIPTION columns in both files (patterns: desc, name, item, product, material …)
  Step 3 — Multi-stage matching:
              Stage 1: your CATEGORY col  vs customer CATEGORY col
              Stage 2: your DESCRIPTION col vs customer DESCRIPTION col
              A pair must pass BOTH stages (score >= threshold) to be kept.
              Final score = average of both stages.
  If only one type of column is found (no category or no description),
  falls back to single-stage match on whichever is available.
  If nothing is detected at all, skips that file with a warning.

EXACT MATCHING (all files):
  Auto-detects SKU / barcode / code columns — no hardcoding.
  Shows which column pair drove each match.

DEDUPLICATION (before ANY matching):
  3-pass: exact rows → cross-sheet → normalised (whitespace/case/punct)
  Removed rows go to "Duplicates Log" sheet.

SCORE REPORTING:
  Every fuzzy row shows Score %.
  Score band histogram printed to console + saved in Summary sheet.

HOW TO USE:
  1. Edit OUR_FILE and CUSTOMER_FILES below.
  2. pip install openpyxl rapidfuzz pandas xlrd
  3. python material_matcher.py

OUTPUT: match_report.xlsx — 6 sheets
  • Summary              – match counts + duplicates + score bands
  • Exact Matches        – code-column matches (shows which col)
  • Fuzzy Matches        – name matches (shows cols, stages, score %)
  • Unmatched Ours       – your items with zero matches
  • Unmatched Customers  – customer items with zero matches
  • Duplicates Log       – every removed duplicate row
"""

import os, re, sys, warnings
from pathlib import Path
import pandas as pd
from rapidfuzz import fuzz, process
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ═══════════════════════════════════════════════════════════════════════
# CONFIGURATION  — only file paths here, zero column names
# ═══════════════════════════════════════════════════════════════════════

OUR_FILE = r"C:\Users\HP\Desktop\functionand sp.csv"

CUSTOMER_FILES = [
    r"C:\Users\HP\Downloads\Al Meera July 2024.xlsx",
    r"C:\Users\HP\Downloads\Talabat March.xlsx",
    r"C:\Users\HP\Downloads\fwqniemtselloutdashboardsession2\C4 sep 24.xlsx",
    r"C:\Users\HP\Downloads\fwqniemtselloutdashboardsession2\Grand Mall APR 2024.xlsx",
    r"C:\Users\HP\Downloads\fwqniemtselloutdashboardsession2\QNIE..xlsx",
    r"C:\Users\HP\Downloads\fwqniemtselloutdashboardsession2\RawSparData.xlsx",
]

OUTPUT_FILE    = r"C:\Users\HP\Desktop\match_report.xlsx"
FUZZY_THRESHOLD = 80   # 0–100. Only scores >= this are kept.

# ═══════════════════════════════════════════════════════════════════════
# COLUMN DETECTION PATTERNS
# Three independent groups — CODE, CATEGORY, DESCRIPTION
# Each group has its own pattern list.
# A column is assigned to the FIRST group whose pattern it matches.
# ═══════════════════════════════════════════════════════════════════════

# ── Group 1: exact / code columns (used for EXACT matching) ───────────
CODE_PATTERNS = [
    r"(?i)\bsku\b",
    r"(?i)\bbarcode\b",
    r"(?i)\bbar[\s_-]?code\b",
    r"(?i)\bupc\b",
    r"(?i)\bean\b",
    r"(?i)\bgtin\b",
    r"(?i)\bitem[\s_-]?(code|no|number|id|#)\b",
    r"(?i)\bproduct[\s_-]?(code|no|number|id|#)\b",
    r"(?i)\bmaterial[\s_-]?(code|no|number|id|#)\b",
    r"(?i)\bpart[\s_-]?(code|no|number|id|#)\b",
    r"(?i)\barticle[\s_-]?(code|no|number|id|#)\b",
    r"(?i)\bcatalog[\s_-]?(code|no|number|id|#)\b",
    r"(?i)\bref(?:erence)?[\s_-]?(code|no|number|id|#)?\b",
    r"(?i)\bhsn\b",
    r"(?i)\bmodel[\s_-]?(code|no|number|id|#)\b",
]

# ── Group 2: CATEGORY columns (used as Stage 1 in fuzzy) ──────────────
# Matches broad grouping / hierarchy columns.
# Add more patterns here if your files use unusual naming.
CATEGORY_PATTERNS = [
    r"(?i)\bcategor",          # category, categorylevel, category_name …
    r"(?i)\bmgrp",             # mgrp, mgrp_descr …
    r"(?i)\bfamily",           # family, familytext, family_name …
    r"(?i)\bgroup\b",          # group, product_group …
    r"(?i)\bdepartment\b",
    r"(?i)\bdept\b",
    r"(?i)\bdivision\b",
    r"(?i)\bsection\b",
    r"(?i)\bclass\b",
    r"(?i)\bsegment\b",
    r"(?i)\bhierarch",         # hierarchy, hierarchy_level …
    r"(?i)\blevel\b",          # level, category level …
    r"(?i)\btype\b",
    r"(?i)\brange\b",
]

# ── Group 3: DESCRIPTION columns (used as Stage 2 in fuzzy) ───────────
# Matches specific item / product name columns.
DESCRIPTION_PATTERNS = [
    r"(?i)\bdescri",           # description, desc, su_description …
    r"(?i)\bproduct[\s_-]?name\b",
    r"(?i)\bitem[\s_-]?name\b",
    r"(?i)\bproduct\b",        # product, productname …
    r"(?i)\bitem\b",
    r"(?i)\bmaterial[\s_-]?desc",
    r"(?i)\bmat[\s_-]?desc",
    r"(?i)\bname\b",
    r"(?i)\btitle\b",
    r"(?i)\bgoods\b",
    r"(?i)\bcommodity\b",
    r"(?i)\blabel\b",
    r"(?i)\bvariant\b",
    r"(?i)\bbrand\b",
]


# ═══════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════

def _match_patterns(col_name, patterns):
    """Return True if col_name matches any pattern in the list."""
    for pat in patterns:
        if re.search(pat, str(col_name).strip()):
            return True
    return False


def detect_columns(df, patterns):
    """Return all columns in df whose name matches any pattern."""
    return [c for c in df.columns if _match_patterns(c, patterns)]


def detect_column_groups(df):
    """
    Detect CODE, CATEGORY, and DESCRIPTION columns in df.
    CODE columns are excluded from the other two groups.
    Returns dict with keys 'code', 'category', 'description'.
    """
    code_cols = detect_columns(df, CODE_PATTERNS)
    code_set  = set(code_cols)

    cat_cols  = [c for c in df.columns
                 if c not in code_set and _match_patterns(c, CATEGORY_PATTERNS)]
    desc_cols = [c for c in df.columns
                 if c not in code_set and _match_patterns(c, DESCRIPTION_PATTERNS)]

    # Remove overlap: if a column matches both CATEGORY and DESCRIPTION,
    # prefer CATEGORY (it's usually the broader grouping).
    cat_set   = set(cat_cols)
    desc_cols = [c for c in desc_cols if c not in cat_set]

    return {"code": code_cols, "category": cat_cols, "description": desc_cols}


def normalize(value):
    if pd.isna(value):
        return ""
    s = str(value).strip().lower()
    s = re.sub(r"[^\w\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def detect_engine(filepath):
    ext = Path(filepath).suffix.lower()
    if ext in (".csv", ".tsv"):
        return "csv", ext
    with open(filepath, "rb") as f:
        header = f.read(8)
    if header[:4] == b"PK\x03\x04":
        return "pyxlsb" if ext == ".xlsb" else "openpyxl", ext
    if header[:4] == b"\xd0\xcf\x11\xe0":
        return "xlrd", ext
    with open(filepath, "rb") as f:
        head = f.read(512).lower()
    if b"<html" in head or b"<table" in head:
        return "html", ext
    return "openpyxl", ext


def combined_text(row, cols):
    """Concatenate normalised values of given cols into one string."""
    return " ".join(normalize(row.get(c, "")) for c in cols if normalize(row.get(c, ""))).strip()


# ═══════════════════════════════════════════════════════════════════════
# DEDUPLICATION
# ═══════════════════════════════════════════════════════════════════════

def dedup_dataframe(df, label):
    """
    3-pass deduplication — runs BEFORE any matching.

    Pass 1 — exact row duplicates (all data cols identical).
              Also catches cross-sheet duplicates because _source_sheet
              is excluded from the comparison set.
    Pass 3 — normalised fingerprint: same content, different
              whitespace / capitalisation / punctuation.

    Returns (clean_df, removed_df).
    """
    data_cols      = [c for c in df.columns if not c.startswith("_")]
    removed_frames = []

    # Pass 1
    before = len(df)
    mask   = df.duplicated(subset=data_cols, keep="first")
    r1     = df[mask].copy()
    r1["_removed_reason"] = "Exact duplicate row"
    r1["_removed_from"]   = label
    removed_frames.append(r1)
    df     = df[~mask].reset_index(drop=True)
    if before - len(df):
        print(f"      Pass 1 (exact):       {before - len(df):,} rows removed")

    # Pass 3 — normalised fingerprint
    fps    = df[data_cols].apply(lambda row: "||".join(normalize(v) for v in row), axis=1)
    before = len(df)
    seen, keep = {}, []
    for i, fp in enumerate(fps):
        if fp in seen:
            keep.append(False)
        else:
            seen[fp] = i
            keep.append(True)
    keep_s = pd.Series(keep, dtype=bool)
    r3     = df[~keep_s].copy()
    r3["_removed_reason"] = "Normalised duplicate (whitespace/case/punctuation)"
    r3["_removed_from"]   = label
    removed_frames.append(r3)
    df     = df[keep_s].reset_index(drop=True)
    if before - len(df):
        print(f"      Pass 3 (normalised):  {before - len(df):,} rows removed")

    removed = pd.concat(removed_frames, ignore_index=True) if removed_frames else pd.DataFrame()
    return df, removed


# ═══════════════════════════════════════════════════════════════════════
# FILE READER
# ═══════════════════════════════════════════════════════════════════════

def read_file(filepath):
    engine, ext = detect_engine(filepath)
    fname       = Path(filepath).name
    frames      = []
    try:
        if engine == "csv":
            df = pd.read_csv(filepath, dtype=str,
                             sep="\t" if ext == ".tsv" else ",")
            df.dropna(how="all", inplace=True)
            df.columns          = [str(c).strip() for c in df.columns]
            df["_source_file"]  = fname
            df["_source_sheet"] = "Sheet1"
            return df
        if engine == "html":
            for i, df in enumerate(pd.read_html(filepath, dtype=str)):
                df.dropna(how="all", inplace=True)
                df.columns          = [str(c).strip() for c in df.columns]
                df["_source_file"]  = fname
                df["_source_sheet"] = f"Table{i+1}"
                frames.append(df)
        else:
            xls = pd.ExcelFile(filepath, engine=engine)
            for sheet in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet, dtype=str)
                df.dropna(how="all", inplace=True)
                df.columns          = [str(c).strip() for c in df.columns]
                df["_source_file"]  = fname
                df["_source_sheet"] = sheet
                frames.append(df)
    except Exception as e:
        print(f"   ⚠️  Error reading {fname}: {e} — trying fallbacks...")
        for fallback in ["openpyxl", "xlrd"]:
            try:
                xls = pd.ExcelFile(filepath, engine=fallback)
                for sheet in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=sheet, dtype=str)
                    df.dropna(how="all", inplace=True)
                    df.columns          = [str(c).strip() for c in df.columns]
                    df["_source_file"]  = fname
                    df["_source_sheet"] = sheet
                    frames.append(df)
                print(f"   ✅  Loaded with '{fallback}'")
                break
            except Exception:
                continue
        else:
            print(f"   ❌  Cannot read {fname}. Skipping.")
            return pd.DataFrame()

    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


# ═══════════════════════════════════════════════════════════════════════
# EXACT MATCHING
# ═══════════════════════════════════════════════════════════════════════

def exact_match(our_df, our_code_cols, cust_df, cust_code_cols):
    """
    Match rows where code values are identical after normalisation.
    Records which column pair drove each match.
    Deduplicates pairs so the same (our, cust) is recorded only once.
    """
    if not our_code_cols or not cust_code_cols:
        return [], set(), set()

    # Build lookup: normalised_value → [(our_idx, our_col), …]
    lookup = {}
    for idx, row in our_df.iterrows():
        for col in our_code_cols:
            val = normalize(row.get(col, ""))
            if val:
                lookup.setdefault(val, []).append((idx, col))

    results, matched_our, matched_cust, seen = [], set(), set(), set()

    for cust_idx, cust_row in cust_df.iterrows():
        for cust_col in cust_code_cols:
            val = normalize(cust_row.get(cust_col, ""))
            if not val or val not in lookup:
                continue
            for our_idx, our_col in lookup[val]:
                pair = (our_idx, cust_idx)
                if pair in seen:
                    continue
                seen.add(pair)
                results.append({
                    "Match Type":     "EXACT",
                    "Our Match Col":  our_col,
                    "Cust Match Col": cust_col,
                    "Matched Value":  val,
                    "Score":          100,
                    "Match Detail":   f"{our_col} ↔ {cust_col} = '{val}'",
                    "our_index":      our_idx,
                    "cust_index":     cust_idx,
                })
                matched_our.add(our_idx)
                matched_cust.add(cust_idx)

    return results, matched_our, matched_cust


# ═══════════════════════════════════════════════════════════════════════
# FUZZY — SINGLE STAGE WORKER
# ═══════════════════════════════════════════════════════════════════════

def _run_stage(our_df, our_cols, cust_df, cust_cols,
               threshold, stage_label,
               restrict_our=None, restrict_cust=None):
    """
    One fuzzy stage: combine our_cols into a text, combine cust_cols
    into a text, then run rapidfuzz token_sort_ratio.

    restrict_our / restrict_cust — index sets to limit the search to
    (used by multi-stage narrowing).

    Returns list of raw hit dicts.
    """
    if not our_cols or not cust_cols:
        return []

    # Validate columns exist
    our_cols_ok  = [c for c in our_cols  if c in our_df.columns]
    cust_cols_ok = [c for c in cust_cols if c in cust_df.columns]
    if not our_cols_ok or not cust_cols_ok:
        missing_our  = [c for c in our_cols  if c not in our_df.columns]
        missing_cust = [c for c in cust_cols if c not in cust_df.columns]
        if missing_our:
            print(f"         ⚠️  Stage [{stage_label}] — "
                  f"cols not found in YOUR file: {missing_our}")
        if missing_cust:
            print(f"         ⚠️  Stage [{stage_label}] — "
                  f"cols not found in customer file: {missing_cust}")
        return []

    our_sub   = our_df  if restrict_our   is None else our_df.loc[list(restrict_our)]
    cust_sub  = cust_df if restrict_cust  is None else cust_df.loc[list(restrict_cust)]

    our_texts = {}
    for idx, row in our_sub.iterrows():
        txt = combined_text(row, our_cols_ok)
        if txt:
            our_texts[idx] = txt

    if not our_texts:
        return []

    our_keys = list(our_texts.keys())
    our_vals = list(our_texts.values())
    results  = []

    for cust_idx, cust_row in cust_sub.iterrows():
        cv = combined_text(cust_row, cust_cols_ok)
        if not cv:
            continue
        hits = process.extract(cv, our_vals,
                               scorer=fuzz.token_sort_ratio,
                               limit=3,
                               score_cutoff=threshold)
        for _, score, pos in hits:
            results.append({
                "our_index":   our_keys[pos],
                "cust_index":  cust_idx,
                "score":       round(score, 1),
                "our_cols":    our_cols_ok,
                "cust_cols":   cust_cols_ok,
                "stage_label": stage_label,
                "our_val":     our_texts[our_keys[pos]],
                "cust_val":    cv,
            })
    return results


# ═══════════════════════════════════════════════════════════════════════
# FUZZY — MAIN ORCHESTRATOR (fully dynamic, works for ALL files)
# ═══════════════════════════════════════════════════════════════════════

def fuzzy_match_for_file(our_df, our_groups,
                          cust_file_df,
                          threshold, skip_our, skip_cust):
    """
    Runs dynamic multi-stage fuzzy matching for ONE customer file.

    Detects column groups in the customer file, then:
      - If BOTH category AND description columns found:
          Stage 1: our CATEGORY cols vs cust CATEGORY cols
          Stage 2: our DESCRIPTION cols vs cust DESCRIPTION cols  (narrows Stage 1 pairs)
          Final score = average of both stages.
          A pair must pass BOTH stages.
      - If only CATEGORY found:
          Single stage on category columns.
      - If only DESCRIPTION found:
          Single stage on description columns.
      - If NEITHER found:
          Skip with warning.

    This logic is IDENTICAL for every file — QNIE, Grand Mall, C4, etc.
    No file-specific rules needed.
    """
    remaining_cust = cust_file_df[~cust_file_df.index.isin(skip_cust)].copy()
    remaining_our  = our_df[~our_df.index.isin(skip_our)].copy()
    fname          = cust_file_df["_source_file"].iloc[0] if len(cust_file_df) else "?"

    # Detect column groups in THIS customer file
    cust_groups = detect_column_groups(cust_file_df)

    our_cat_cols   = our_groups["category"]
    our_desc_cols  = our_groups["description"]
    cust_cat_cols  = cust_groups["category"]
    cust_desc_cols = cust_groups["description"]

    print(f"      Detected columns:")
    print(f"         Your  category:    {our_cat_cols  or '(none)'}")
    print(f"         Your  description: {our_desc_cols or '(none)'}")
    print(f"         Cust  category:    {cust_cat_cols  or '(none)'}")
    print(f"         Cust  description: {cust_desc_cols or '(none)'}")

    has_cat  = bool(our_cat_cols  and cust_cat_cols)
    has_desc = bool(our_desc_cols and cust_desc_cols)

    if not has_cat and not has_desc:
        print(f"      ⚠️  No matchable columns found — skipping {fname}")
        return [], set(), set()

    results, matched_our, matched_cust, seen_pairs = [], set(), set(), set()

    # ── Two-stage: category → description ─────────────────────────────
    if has_cat and has_desc:
        print(f"      Mode: 2-stage  [Category → Description]")

        # Stage 1 — category
        print(f"      Stage 1 [Category]: "
              f"{our_cat_cols} vs {cust_cat_cols}...")
        stage1_hits = _run_stage(remaining_our, our_cat_cols,
                                  remaining_cust, cust_cat_cols,
                                  threshold, "Category")
        print(f"         {len(stage1_hits):,} candidates")

        if stage1_hits:
            # Stage 2 — description, restricted to Stage 1 survivors
            cand_our  = {r["our_index"]  for r in stage1_hits}
            cand_cust = {r["cust_index"] for r in stage1_hits}
            print(f"      Stage 2 [Description]: "
                  f"{our_desc_cols} vs {cust_desc_cols}  "
                  f"(narrowing {len(cand_cust):,} cust rows)...")
            stage2_hits = _run_stage(remaining_our, our_desc_cols,
                                      remaining_cust, cust_desc_cols,
                                      threshold, "Description",
                                      restrict_our=cand_our,
                                      restrict_cust=cand_cust)
            print(f"         {len(stage2_hits):,} still passing")

            # Keep only pairs that appear in BOTH stages
            s1_map = {(r["our_index"], r["cust_index"]): r for r in stage1_hits}
            s2_map = {(r["our_index"], r["cust_index"]): r for r in stage2_hits}
            common = s1_map.keys() & s2_map.keys()
            print(f"         {len(common):,} pairs passed both stages")

            for pair in common:
                if pair in seen_pairs:
                    continue
                seen_pairs.add(pair)
                p1, p2  = s1_map[pair], s2_map[pair]
                avg_score = round((p1["score"] + p2["score"]) / 2, 1)
                col_our   = list(dict.fromkeys(p1["our_cols"]  + p2["our_cols"]))
                col_cust  = list(dict.fromkeys(p1["cust_cols"] + p2["cust_cols"]))
                results.append({
                    "Match Type":     "FUZZY",
                    "Our Match Col":  ", ".join(col_our),
                    "Cust Match Col": ", ".join(col_cust),
                    "Match Stages":   "Category → Description",
                    "Our Value":      f"{p1['our_val']} | {p2['our_val']}",
                    "Cust Value":     f"{p1['cust_val']} | {p2['cust_val']}",
                    "Score":          avg_score,
                    "Match Detail":   (f"[Cat→Desc]  "
                                       f"{p1['our_val']}|{p2['our_val']}  ≈  "
                                       f"{p1['cust_val']}|{p2['cust_val']}  "
                                       f"({avg_score}%)"),
                    "our_index":      pair[0],
                    "cust_index":     pair[1],
                })
                matched_our.add(pair[0])
                matched_cust.add(pair[1])

    # ── Single-stage: category only ────────────────────────────────────
    elif has_cat:
        print(f"      Mode: 1-stage  [Category only]")
        print(f"      {our_cat_cols} vs {cust_cat_cols}...")
        hits = _run_stage(remaining_our, our_cat_cols,
                           remaining_cust, cust_cat_cols,
                           threshold, "Category")
        print(f"         {len(hits):,} matches")
        for r in hits:
            pair = (r["our_index"], r["cust_index"])
            if pair in seen_pairs:
                continue
            seen_pairs.add(pair)
            results.append({
                "Match Type":     "FUZZY",
                "Our Match Col":  ", ".join(r["our_cols"]),
                "Cust Match Col": ", ".join(r["cust_cols"]),
                "Match Stages":   "Category",
                "Our Value":      r["our_val"],
                "Cust Value":     r["cust_val"],
                "Score":          r["score"],
                "Match Detail":   (f"[Category]  {r['our_val']}  ≈  "
                                   f"{r['cust_val']}  ({r['score']}%)"),
                "our_index":      r["our_index"],
                "cust_index":     r["cust_index"],
            })
            matched_our.add(r["our_index"])
            matched_cust.add(r["cust_index"])

    # ── Single-stage: description only ────────────────────────────────
    else:
        print(f"      Mode: 1-stage  [Description only]")
        print(f"      {our_desc_cols} vs {cust_desc_cols}...")
        total = len(remaining_cust)
        hits  = []
        our_texts = {}
        our_desc_ok = [c for c in our_desc_cols if c in remaining_our.columns]
        for idx, row in remaining_our.iterrows():
            txt = combined_text(row, our_desc_ok)
            if txt:
                our_texts[idx] = txt
        our_keys = list(our_texts.keys())
        our_vals = list(our_texts.values())
        cust_desc_ok = [c for c in cust_desc_cols if c in remaining_cust.columns]
        for i, (cust_idx, cust_row) in enumerate(remaining_cust.iterrows()):
            if i % 500 == 0:
                print(f"         {i}/{total}...", end="\r")
            cv = combined_text(cust_row, cust_desc_ok)
            if not cv:
                continue
            raw = process.extract(cv, our_vals,
                                  scorer=fuzz.token_sort_ratio,
                                  limit=3, score_cutoff=threshold)
            for _, score, pos in raw:
                hits.append({
                    "our_index":  our_keys[pos],
                    "cust_index": cust_idx,
                    "score":      round(score, 1),
                    "our_cols":   our_desc_ok,
                    "cust_cols":  cust_desc_ok,
                    "our_val":    our_texts[our_keys[pos]],
                    "cust_val":   cv,
                })
        print(f"\n         {len(hits):,} matches")
        for r in hits:
            pair = (r["our_index"], r["cust_index"])
            if pair in seen_pairs:
                continue
            seen_pairs.add(pair)
            results.append({
                "Match Type":     "FUZZY",
                "Our Match Col":  ", ".join(r["our_cols"]),
                "Cust Match Col": ", ".join(r["cust_cols"]),
                "Match Stages":   "Description",
                "Our Value":      r["our_val"],
                "Cust Value":     r["cust_val"],
                "Score":          r["score"],
                "Match Detail":   (f"[Description]  {r['our_val']}  ≈  "
                                   f"{r['cust_val']}  ({r['score']}%)"),
                "our_index":      r["our_index"],
                "cust_index":     r["cust_index"],
            })
            matched_our.add(r["our_index"])
            matched_cust.add(r["cust_index"])

    return results, matched_our, matched_cust


# ═══════════════════════════════════════════════════════════════════════
# SCORE DISTRIBUTION
# ═══════════════════════════════════════════════════════════════════════

def score_distribution(fuzzy_results):
    bands = {"80–84%": 0, "85–89%": 0, "90–94%": 0, "95–99%": 0, "100%": 0}
    for r in fuzzy_results:
        s = r["Score"]
        if   s == 100: bands["100%"]   += 1
        elif s >= 95:  bands["95–99%"] += 1
        elif s >= 90:  bands["90–94%"] += 1
        elif s >= 85:  bands["85–89%"] += 1
        else:          bands["80–84%"] += 1
    return bands


# ═══════════════════════════════════════════════════════════════════════
# REPORT
# ═══════════════════════════════════════════════════════════════════════

HDR_COLOR       = "1F4E79"
EXACT_COLOR     = "C6EFCE"
FUZZY_COLOR     = "FFEB9C"
UNMATCHED_COLOR = "FCE4D6"
DUPES_COLOR     = "EDEDED"


def build_match_rows(matches, our_df, cust_df):
    rows = []
    for m in matches:
        our_row  = our_df.loc[m["our_index"]]
        cust_row = cust_df.loc[m["cust_index"]]
        row = {
            "Match Type":       m.get("Match Type"),
            "Score %":          m.get("Score"),
            "Match Detail":     m.get("Match Detail"),
            "Our Col(s) Used":  m.get("Our Match Col"),
            "Cust Col(s) Used": m.get("Cust Match Col"),
            "Customer File":    cust_row.get("_source_file", ""),
            "Customer Sheet":   cust_row.get("_source_sheet", ""),
        }
        if m.get("Match Type") == "FUZZY":
            row["Match Stages"] = m.get("Match Stages", "")
            row["Our Value"]    = m.get("Our Value", "")
            row["Cust Value"]   = m.get("Cust Value", "")
        for c in our_df.columns:
            if not c.startswith("_"):
                row[f"OUR_{c}"] = our_row.get(c, "")
        for c in cust_df.columns:
            if not c.startswith("_"):
                row[f"CUST_{c}"] = cust_row.get(c, "")
        rows.append(row)
    return pd.DataFrame(rows)


def format_sheet(ws, header_hex, row_hex=None):
    hdr_fill  = PatternFill("solid", fgColor=header_hex)
    hdr_font  = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = hdr_align
    if row_hex and ws.max_row > 1:
        fill = PatternFill("solid", fgColor=row_hex)
        font = Font(name="Calibri", size=10)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.fill = fill
                cell.font = font
    for col_cells in ws.iter_cols(max_row=min(200, ws.max_row)):
        w = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(w + 2, 45)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes    = "A2"


def write_report(exact_df, fuzzy_df, unmatched_our, unmatched_cust,
                 our_df, cust_df, all_removed_df, score_bands, path):
    print("\n📝 Writing report...")

    rows = []
    for fname in cust_df["_source_file"].unique():
        sub       = cust_df[cust_df["_source_file"] == fname]
        n_total   = len(sub)
        n_exact   = len(exact_df[exact_df["Customer File"] == fname]) if not exact_df.empty else 0
        n_fuzzy   = len(fuzzy_df[fuzzy_df["Customer File"] == fname]) if not fuzzy_df.empty else 0
        n_matched = n_exact + n_fuzzy
        n_removed = len(all_removed_df[all_removed_df["_removed_from"] == fname]) \
                    if not all_removed_df.empty else 0
        rows.append({
            "Customer File":      fname,
            "Rows (after dedup)": n_total,
            "Duplicates Removed": n_removed,
            "Exact Matches":      n_exact,
            "Fuzzy Matches":      n_fuzzy,
            "Total Matched":      n_matched,
            "Unmatched":          n_total - n_matched,
            "Match Rate %":       round(100 * n_matched / n_total, 1) if n_total else 0,
        })
    summary_df  = pd.DataFrame(rows)
    total_rows  = summary_df["Rows (after dedup)"].sum()
    summary_df  = pd.concat([summary_df, pd.DataFrame([{
        "Customer File":      "TOTAL",
        "Rows (after dedup)": total_rows,
        "Duplicates Removed": summary_df["Duplicates Removed"].sum(),
        "Exact Matches":      summary_df["Exact Matches"].sum(),
        "Fuzzy Matches":      summary_df["Fuzzy Matches"].sum(),
        "Total Matched":      summary_df["Total Matched"].sum(),
        "Unmatched":          summary_df["Unmatched"].sum(),
        "Match Rate %":       round(100 * summary_df["Total Matched"].sum() / total_rows, 1)
                              if total_rows else 0,
    }])], ignore_index=True)

    # Score bands appended to summary
    summary_df = pd.concat([summary_df, pd.DataFrame(
        [{"Customer File": "── Fuzzy Score Distribution ──"}] +
        [{"Customer File": f"  {b}", "Fuzzy Matches": c} for b, c in score_bands.items()]
    )], ignore_index=True)

    drop_our  = [c for c in our_df.columns  if c.startswith("_")]
    drop_cust = [c for c in cust_df.columns if c.startswith("_")]

    if not all_removed_df.empty:
        mc   = ["_removed_reason", "_removed_from"]
        rest = [c for c in all_removed_df.columns if c not in mc]
        dupes_out = all_removed_df[mc + rest].rename(
            columns={"_removed_reason": "Reason Removed",
                     "_removed_from":   "Removed From"})
    else:
        dupes_out = pd.DataFrame([{"Info": "No duplicates found"}])

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        (exact_df.drop(columns=["our_index","cust_index"], errors="ignore")
         if not exact_df.empty
         else pd.DataFrame([{"Info": "No exact matches found"}])
        ).to_excel(writer, sheet_name="Exact Matches", index=False)
        (fuzzy_df.drop(columns=["our_index","cust_index"], errors="ignore")
         if not fuzzy_df.empty
         else pd.DataFrame([{"Info": "No fuzzy matches found"}])
        ).to_excel(writer, sheet_name="Fuzzy Matches", index=False)
        unmatched_our.drop(columns=drop_our,   errors="ignore").to_excel(
            writer, sheet_name="Unmatched - Ours",      index=False)
        unmatched_cust.drop(columns=drop_cust, errors="ignore").to_excel(
            writer, sheet_name="Unmatched - Customers", index=False)
        dupes_out.to_excel(writer, sheet_name="Duplicates Log", index=False)

    wb = load_workbook(path)
    format_sheet(wb["Summary"],               HDR_COLOR)
    format_sheet(wb["Exact Matches"],         HDR_COLOR, EXACT_COLOR)
    format_sheet(wb["Fuzzy Matches"],         HDR_COLOR, FUZZY_COLOR)
    format_sheet(wb["Unmatched - Ours"],      HDR_COLOR, UNMATCHED_COLOR)
    format_sheet(wb["Unmatched - Customers"], HDR_COLOR, UNMATCHED_COLOR)
    format_sheet(wb["Duplicates Log"],        HDR_COLOR, DUPES_COLOR)
    wb.save(path)
    print(f"✅  Report saved → {path}")


# ═══════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════

def main():
    print("=" * 65)
    print("  MATERIAL MATCHER — Exact + Dynamic Fuzzy Matching")
    print("=" * 65)

    if not os.path.exists(OUR_FILE):
        print(f"\n❌ Our file not found: {OUR_FILE}")
        sys.exit(1)

    all_removed = []

    # ── Load + dedup OUR file ──────────────────────────────────────────
    print(f"\n📂 Loading our file...")
    our_df = read_file(OUR_FILE)
    print(f"   → {len(our_df):,} rows")
    our_df, our_rem = dedup_dataframe(our_df, "OUR FILE")
    print(f"   ✅ {len(our_df):,} kept  ({len(our_rem):,} duplicates removed)")
    if not our_rem.empty:
        all_removed.append(our_rem)

    # Detect column groups in OUR file once — reused for every customer file
    our_groups = detect_column_groups(our_df)
    print(f"\n   Your file column groups detected:")
    print(f"      Code cols:        {our_groups['code']        or '(none)'}")
    print(f"      Category cols:    {our_groups['category']    or '(none)'}")
    print(f"      Description cols: {our_groups['description'] or '(none)'}")

    # ── Load + dedup customer files ────────────────────────────────────
    cust_frames = []
    for f in CUSTOMER_FILES:
        if not os.path.exists(f):
            print(f"\n   ⚠️  Missing: {f}")
            continue
        fname = Path(f).name
        print(f"\n📂 Loading: {fname}")
        df = read_file(f)
        if df.empty:
            continue
        print(f"   → {len(df):,} rows")
        df, rem = dedup_dataframe(df, fname)
        print(f"   ✅ {len(df):,} kept  ({len(rem):,} duplicates removed)")
        if not rem.empty:
            all_removed.append(rem)
        cust_frames.append(df)

    if not cust_frames:
        print("\n❌ No customer files loaded.")
        sys.exit(1)

    cust_df = pd.concat(cust_frames, ignore_index=True)
    print(f"\n📊 All customer rows (post-dedup): {len(cust_df):,}")
    print(f"   🔍 Cross-file deduplication...")
    cust_df, cross_rem = dedup_dataframe(cust_df, "CROSS-FILE")
    if not cross_rem.empty:
        print(f"   ✅ {len(cross_rem):,} cross-file duplicates removed")
        all_removed.append(cross_rem)
    else:
        print(f"   ✅ No cross-file duplicates found")

    all_removed_df = pd.concat(all_removed, ignore_index=True) \
                     if all_removed else pd.DataFrame()

    # ── EXACT MATCHING ─────────────────────────────────────────────────
    cust_code_cols = detect_columns(cust_df, CODE_PATTERNS)
    print(f"\n⚡ Exact matching:")
    print(f"   Your code cols: {our_groups['code']  or '(none)'}")
    print(f"   Cust code cols: {cust_code_cols or '(none)'}")
    exact_results, m_our_e, m_cust_e = exact_match(
        our_df, our_groups["code"], cust_df, cust_code_cols)
    print(f"   → {len(exact_results):,} exact matches")
    if exact_results:
        col_counts = {}
        for r in exact_results:
            k = f"{r['Our Match Col']} ↔ {r['Cust Match Col']}"
            col_counts[k] = col_counts.get(k, 0) + 1
        for k, v in sorted(col_counts.items(), key=lambda x: -x[1]):
            print(f"      {k}: {v:,}")

    # ── FUZZY MATCHING — per file ──────────────────────────────────────
    print(f"\n🔄 Fuzzy matching (threshold ≥ {FUZZY_THRESHOLD}%)...")
    all_fuzzy = []
    m_our_f, m_cust_f = set(), set()

    for fname in cust_df["_source_file"].unique():
        file_sub = cust_df[cust_df["_source_file"] == fname]
        print(f"\n   📄 {fname}  ({len(file_sub):,} rows)")

        f_res, f_our, f_cust = fuzzy_match_for_file(
            our_df, our_groups,
            file_sub,
            FUZZY_THRESHOLD,
            skip_our  = m_our_e,
            skip_cust = m_cust_e | m_cust_f,
        )
        print(f"      → {len(f_res):,} fuzzy matches kept (score ≥ {FUZZY_THRESHOLD}%)")
        all_fuzzy.extend(f_res)
        m_our_f  |= f_our
        m_cust_f |= f_cust

    # ── Score distribution ─────────────────────────────────────────────
    score_bands = score_distribution(all_fuzzy)

    # ── Assemble output ────────────────────────────────────────────────
    all_matched_our  = m_our_e | m_our_f
    all_matched_cust = m_cust_e | m_cust_f

    exact_df_out   = build_match_rows(exact_results, our_df, cust_df)
    fuzzy_df_out   = build_match_rows(all_fuzzy,     our_df, cust_df)
    unmatched_our  = our_df.loc[~our_df.index.isin(all_matched_our)]
    unmatched_cust = cust_df.loc[~cust_df.index.isin(all_matched_cust)]

    # ── Console summary ────────────────────────────────────────────────
    print(f"\n{'=' * 65}")
    print(f"  FINAL SUMMARY")
    print(f"{'=' * 65}")
    print(f"  Your items (after dedup):       {len(our_df):,}")
    print(f"  Customer items (after dedup):   {len(cust_df):,}")
    print(f"  Total duplicates removed:       {len(all_removed_df):,}")
    print(f"  Exact matches:                  {len(exact_results):,}")
    print(f"  Fuzzy matches:                  {len(all_fuzzy):,}")
    print(f"  Unmatched (ours):               {len(unmatched_our):,}")
    print(f"  Unmatched (customers):          {len(unmatched_cust):,}")
    print(f"\n  Fuzzy Score Bands (>= {FUZZY_THRESHOLD}%):")
    for band, count in score_bands.items():
        bar = "█" * min(count, 50)
        print(f"    {band:>8}  {bar}  {count:,}")

    write_report(exact_df_out, fuzzy_df_out, unmatched_our, unmatched_cust,
                 our_df, cust_df, all_removed_df, score_bands, OUTPUT_FILE)


if __name__ == "__main__":
    main()