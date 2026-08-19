"""
Material Matcher — Exact + Optimised Fuzzy Matching
Fully dynamic column detection for all 7 files.

YOUR FILE columns used:
  Exact  : barcode, material_code, old_code
  Fuzzy  : material_desc, mgrp_descr, barcode_description,
            prdh_descr_1, prdh_descr_2, prdh_descr_3, emgrp_desc

CUSTOMER FILES — auto-detected per file:
  Grand Mall   : Product Code → exact | Item Name, Product Group → fuzzy
  C4           : BARCODE → exact | SU_DESCRIPTION, CATEGORY_NAMNE, SUB_CATEG_NAMNE → fuzzy
  RawSparData  : (no barcode) | Family Text, Description, Retail Article Brand Description Text → fuzzy
  Talabat      : Sku, Barcode → exact | Product_Name, L1, L2 → fuzzy
  Al Meera     : Item barcode, Item number → exact | Product name, Category, Brand → fuzzy
  QNIE         : Material → exact | Brand, Material Group → fuzzy

HOW FUZZY WORKS:
  1. ALL name/desc columns on each side are combined into one text string per row.
  2. rapidfuzz token_sort_ratio compares your combined text vs customer combined text.
  3. Category bucket optimization: group your items by mgrp_descr first,
     then only fuzzy-match within matching category buckets (FAST).
     Falls back to full scan if no category column found AND file < 20k rows.
  4. Score >= FUZZY_THRESHOLD to keep a match.
"""

import os, re, sys, warnings, time
from pathlib import Path
import pandas as pd
from rapidfuzz import fuzz, process
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ═══════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════

OUR_FILE = r"C:\Users\HP\Desktop\functionand sp.csv"

CUSTOMER_FILES = [
    r"C:\Users\HP\Downloads\Al Meera July 2024.xlsx",
    r"C:\Users\HP\Downloads\Talabat March.xlsx",
    r"C:\Users\HP\Downloads\fwqniemtselloutdashboardsession2\C4 sep 24.xlsx",
    r"C:\Users\HP\Downloads\fwqniemtselloutdashboardsession2\Grand Mall APR 2024.xlsx",
    r"C:\Users\HP\Downloads\fwqniemtselloutdashboardsession2\QNIE..xlsx",
    r"C:\Users\HP\Downloads\fwqniemtselloutdashboardsession2\RawSparData.xlsx",
]

OUTPUT_FILE     = r"C:\Users\HP\Desktop\match_report.xlsx"
FUZZY_THRESHOLD = 80   # 0-100. Only scores >= this are kept.

# ═══════════════════════════════════════════════════════════════════════
# YOUR FILE — PINNED COLUMN DEFINITIONS
# ═══════════════════════════════════════════════════════════════════════

OUR_EXACT_COLS = ["barcode", "material_code", "old_code"]
OUR_FUZZY_COLS = [
    "material_desc",       # primary product description
    "mgrp_descr",          # material group description (category)
    "barcode_description", # barcode description
    "prdh_descr_1",        # product hierarchy level 1
    "prdh_descr_2",        # product hierarchy level 2
    "prdh_descr_3",        # product hierarchy level 3
    "emgrp_desc",          # extended material group description
]
OUR_CAT_COL = "mgrp_descr"  # used as category bucket key

# ═══════════════════════════════════════════════════════════════════════
# CUSTOMER FILE — DYNAMIC COLUMN DETECTION PATTERNS
# ═══════════════════════════════════════════════════════════════════════

# Columns matching CODE_PATTERNS  → EXACT matching
CODE_PATTERNS = [
    r"(?i)^sku$",
    r"(?i)\bbarcode\b",
    r"(?i)\bbar[\s_-]?code\b",
    r"(?i)\bupc\b", r"(?i)\bean\b", r"(?i)\bgtin\b",
    r"(?i)\bitem[\s_-]?(barcode|code|no|number|id)\b",
    r"(?i)\bproduct[\s_-]?(code|no|number|id)\b",
    r"(?i)^material$",           # QNIE: "Material" col = material code
    r"(?i)\bpart[\s_-]?(code|no|number|id)\b",
    r"(?i)\barticle[\s_-]?(code|no|number|id)\b",
    r"(?i)\bref(?:erence)?[\s_-]?(code|no|number|id)?\b",
]

# Columns matching NAME_PATTERNS  → combined FUZZY text (all used together)
NAME_PATTERNS = [
    r"(?i)\bdescri",              # description, SU_DESCRIPTION, desc
    r"(?i)\bproduct[\s_-]?name\b",
    r"(?i)\bitem[\s_-]?name\b",
    r"(?i)\bname\b",
    r"(?i)\btitle\b",
    r"(?i)\bbrand\b",             # Brand adds brand name to fuzzy text
    r"(?i)\bfamily[\s_-]?text\b",
    r"(?i)\bretail[\s_-]?article",
    r"(?i)\bproduct[\s_-]?group\b",
    r"(?i)\bsub[\s_-]?categ",    # SUB_CATEG_NAMNE
    r"(?i)\bmaterial[\s_-]?group\b",  # QNIE: Material Group
    r"(?i)^l1$", r"(?i)^l2$",   # Talabat hierarchy levels
    r"(?i)\banalysis[\s_-]?level\b",
    r"(?i)\bcategory[\s_-]?level\b",
    r"(?i)\bcategor",            # Category, CATEGORY_NAMNE
]

# Best single column to use as category bucket key per customer file
CAT_PATTERNS = [
    r"(?i)^category$",
    r"(?i)^product[\s_-]?group$",
    r"(?i)^family[\s_-]?text$",
    r"(?i)^l1$",
    r"(?i)^material[\s_-]?group$",
    r"(?i)^category[\s_-]?level",
    r"(?i)^analysis[\s_-]?level",
    r"(?i)^categor",
    r"(?i)^family\b",
    r"(?i)^group\b",
    r"(?i)^department\b",
    r"(?i)^section\b",
]

# Always exclude these from fuzzy matching (numeric/date/operational cols)
EXCLUDE_PATTERNS = [
    r"(?i)\bdate\b", r"(?i)\bmonth\b", r"(?i)\byear\b",
    r"(?i)\bprice\b", r"(?i)\bamount\b", r"(?i)\bvalue\b",
    r"(?i)\bsales\b", r"(?i)\bqty\b", r"(?i)\bquantit",
    r"(?i)\bunit\b", r"(?i)\bstore[\s_-]?name\b",
    r"(?i)\bcogs\b", r"(?i)\bweight\b", r"(?i)\bplant\b",
    r"(?i)\bfrom[\s_-]?date\b", r"(?i)\bto[\s_-]?date\b",
    r"(?i)\bexport[\s_-]?date\b", r"(?i)\bcreated\b",
    r"(?i)\bcountry\b", r"(?i)\bcoo\b",
    r"(?i)\bhs[\s_-]?code\b",
    r"(?i)\bimp[\s_-]?local\b", r"(?i)\bstock[\s_-]?unit\b",
    r"(?i)\bsum\b", r"(?i)\baverage\b",
    r"(?i)\bfulfill",
]

# ═══════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════

def matches_any(col_name, patterns):
    return any(re.search(p, str(col_name).strip()) for p in patterns)

def detect_cols(df, include_patterns, exclude_patterns=None):
    result = []
    for c in df.columns:
        if c.startswith("_"):
            continue
        if exclude_patterns and matches_any(c, exclude_patterns):
            continue
        if matches_any(c, include_patterns):
            result.append(c)
    return result

def pick_cat_col(df):
    for pat in CAT_PATTERNS:
        for c in df.columns:
            if re.search(pat, str(c).strip()) and not matches_any(c, EXCLUDE_PATTERNS):
                return c
    return None

def normalize(value):
    if pd.isna(value):
        return ""
    s = str(value).strip().lower()
    s = re.sub(r"[^\w\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def combined_text(row, cols):
    parts = [normalize(row.get(c, "")) for c in cols]
    return " ".join(p for p in parts if p).strip()

def detect_engine(filepath):
    ext = Path(filepath).suffix.lower()
    if ext in (".csv", ".tsv"):
        return "csv", ext
    with open(filepath, "rb") as f:
        header = f.read(8)
    if header[:4] == b"PK\x03\x04":
        return "pyxlsb" if ext == ".xlsb" else "openpyxl", ext
    if header[:4] == b"\xd0\xcf\x11\xe0":
        return "xlrd", ext
    with open(filepath, "rb") as f:
        head = f.read(512).lower()
    if b"<html" in head or b"<table" in head:
        return "html", ext
    return "openpyxl", ext

def read_file(filepath):
    engine, ext = detect_engine(filepath)
    fname  = Path(filepath).name
    frames = []
    try:
        if engine == "csv":
            df = pd.read_csv(filepath, dtype=str,
                             sep="\t" if ext == ".tsv" else ",")
            df.dropna(how="all", inplace=True)
            df.columns = [str(c).strip() for c in df.columns]
            df["_source_file"]  = fname
            df["_source_sheet"] = "Sheet1"
            return df
        if engine == "html":
            for i, df in enumerate(pd.read_html(filepath, dtype=str)):
                df.dropna(how="all", inplace=True)
                df.columns = [str(c).strip() for c in df.columns]
                df["_source_file"]  = fname
                df["_source_sheet"] = f"Table{i+1}"
                frames.append(df)
        else:
            xls = pd.ExcelFile(filepath, engine=engine)
            for sheet in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet, dtype=str)
                df.dropna(how="all", inplace=True)
                df.columns = [str(c).strip() for c in df.columns]
                df["_source_file"]  = fname
                df["_source_sheet"] = sheet
                frames.append(df)
    except Exception as e:
        print(f"   Warning: {fname}: {e} — trying fallbacks...")
        for fallback in ["openpyxl", "xlrd"]:
            try:
                xls = pd.ExcelFile(filepath, engine=fallback)
                for sheet in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=sheet, dtype=str)
                    df.dropna(how="all", inplace=True)
                    df.columns = [str(c).strip() for c in df.columns]
                    df["_source_file"]  = fname
                    df["_source_sheet"] = sheet
                    frames.append(df)
                print(f"   Loaded with '{fallback}'")
                break
            except Exception:
                continue
        else:
            print(f"   Cannot read {fname}. Skipping.")
            return pd.DataFrame()
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)

# ═══════════════════════════════════════════════════════════════════════
# EXACT MATCHING
# ═══════════════════════════════════════════════════════════════════════

def exact_match(our_df, our_exact_cols, cust_df, cust_code_cols):
    if not our_exact_cols or not cust_code_cols:
        return [], set(), set()

    lookup = {}
    for idx, row in our_df.iterrows():
        for col in our_exact_cols:
            val = normalize(row.get(col, ""))
            if val:
                lookup.setdefault(val, []).append((idx, col))

    results, matched_our, matched_cust, seen = [], set(), set(), set()
    for cust_idx, cust_row in cust_df.iterrows():
        for cust_col in cust_code_cols:
            val = normalize(cust_row.get(cust_col, ""))
            if not val or val not in lookup:
                continue
            for our_idx, our_col in lookup[val]:
                pair = (our_idx, cust_idx)
                if pair in seen:
                    continue
                seen.add(pair)
                results.append({
                    "Match Type":     "EXACT",
                    "Our Match Col":  our_col,
                    "Cust Match Col": cust_col,
                    "Matched Value":  val,
                    "Score":          100,
                    "our_index":      our_idx,
                    "cust_index":     cust_idx,
                })
                matched_our.add(our_idx)
                matched_cust.add(cust_idx)
    return results, matched_our, matched_cust

# ═══════════════════════════════════════════════════════════════════════
# FUZZY MATCHING
# ═══════════════════════════════════════════════════════════════════════

def fuzzy_match_file(our_df, our_fuzzy_cols, our_cat_col,
                     cust_df, cust_name_cols, cust_cat_col,
                     fname, threshold, skip_our, skip_cust):
    """
    Fuzzy match ONE customer file against our catalog.

    Strategy A — Category bucket (FAST, used when cat col found):
      Build index: norm(our_mgrp_descr) → {our_idx: combined_all_fuzzy_text}
      Per customer row: lookup category → get small bucket → fuzzy match within bucket only.
      Reduces 15,809 comparisons to ~10-200 per row.

    Strategy B — Full scan (only for files < 20k rows with no cat col):
      Compare every customer row vs all our rows.
      Safe for small files (RawSparData, QNIE, Grand Mall).
    """
    our_sub  = our_df[~our_df.index.isin(skip_our)].copy()
    cust_sub = cust_df[~cust_df.index.isin(skip_cust)].copy()

    if our_sub.empty or cust_sub.empty:
        print(f"      Nothing left to match — skipping.")
        return [], set(), set()

    our_cols_ok  = [c for c in our_fuzzy_cols  if c in our_df.columns]
    cust_cols_ok = [c for c in cust_name_cols  if c in cust_df.columns]

    if not our_cols_ok or not cust_cols_ok:
        print(f"      No usable text columns — skipping.")
        return [], set(), set()

    has_cat = bool(our_cat_col and our_cat_col in our_df.columns and cust_cat_col)
    total   = len(cust_sub)

    print(f"      Your fuzzy cols : {our_cols_ok}")
    print(f"      Cust fuzzy cols : {cust_cols_ok}")
    print(f"      Cat bucket      : {our_cat_col} ↔ {cust_cat_col or '(none)'}  "
          f"{'✅ FAST' if has_cat else '⚠️  Full scan'}")

    results, matched_our, matched_cust, seen = [], set(), set(), set()
    t0   = time.time()
    hits = 0

    # ── STRATEGY A: Category bucket ────────────────────────────────────
    if has_cat:
        # Build: norm_category → {our_idx: combined_text_of_all_fuzzy_cols}
        cat_index = {}
        for idx, row in our_sub.iterrows():
            cat = normalize(row.get(our_cat_col, ""))
            txt = combined_text(row, our_cols_ok)
            if cat and txt:
                cat_index.setdefault(cat, {})[idx] = txt

        for i, (cust_idx, cust_row) in enumerate(cust_sub.iterrows()):
            if i % 5000 == 0 and i > 0:
                elapsed = time.time() - t0
                rate    = i / elapsed if elapsed > 0 else 1
                eta     = (total - i) / rate
                print(f"         {i:,}/{total:,}  |  {hits:,} hits  |  "
                      f"ETA {eta/60:.1f} min    ", end="\r")

            cust_cat = normalize(cust_row.get(cust_cat_col, ""))
            cust_txt = combined_text(cust_row, cust_cols_ok)
            if not cust_cat or not cust_txt:
                continue

            bucket = cat_index.get(cust_cat, {})
            if not bucket:
                continue

            our_keys = list(bucket.keys())
            our_vals = list(bucket.values())
            raw = process.extract(cust_txt, our_vals,
                                  scorer=fuzz.token_sort_ratio,
                                  limit=3, score_cutoff=threshold)
            for _, score, pos in raw:
                our_idx = our_keys[pos]
                pair    = (our_idx, cust_idx)
                if pair in seen:
                    continue
                seen.add(pair)
                results.append({
                    "Match Type":       "FUZZY",
                    "Our Match Col":    ", ".join(our_cols_ok),
                    "Cust Match Col":   ", ".join(cust_cols_ok),
                    "Category Matched": cust_cat,
                    "Our Text":         our_vals[pos],
                    "Cust Text":        cust_txt,
                    "Score":            round(score, 1),
                    "our_index":        our_idx,
                    "cust_index":       cust_idx,
                })
                matched_our.add(our_idx)
                matched_cust.add(cust_idx)
                hits += 1

    # ── STRATEGY B: Full scan (small files only) ───────────────────────
    else:
        if total > 20000:
            print(f"      No category col found & {total:,} rows — too large for full scan.")
            print(f"      Add a pattern to CAT_PATTERNS to enable bucket matching.")
            return [], set(), set()

        our_texts = {idx: combined_text(row, our_cols_ok)
                     for idx, row in our_sub.iterrows()
                     if combined_text(row, our_cols_ok)}
        our_keys = list(our_texts.keys())
        our_vals = list(our_texts.values())

        for i, (cust_idx, cust_row) in enumerate(cust_sub.iterrows()):
            if i % 500 == 0 and i > 0:
                elapsed = time.time() - t0
                rate    = i / elapsed if elapsed > 0 else 1
                eta     = (total - i) / rate
                print(f"         {i:,}/{total:,}  |  {hits:,} hits  |  "
                      f"ETA {eta/60:.1f} min    ", end="\r")
            cust_txt = combined_text(cust_row, cust_cols_ok)
            if not cust_txt:
                continue
            raw = process.extract(cust_txt, our_vals,
                                  scorer=fuzz.token_sort_ratio,
                                  limit=3, score_cutoff=threshold)
            for _, score, pos in raw:
                our_idx = our_keys[pos]
                pair    = (our_idx, cust_idx)
                if pair in seen:
                    continue
                seen.add(pair)
                results.append({
                    "Match Type":     "FUZZY",
                    "Our Match Col":  ", ".join(our_cols_ok),
                    "Cust Match Col": ", ".join(cust_cols_ok),
                    "Our Text":       our_vals[pos],
                    "Cust Text":      cust_txt,
                    "Score":          round(score, 1),
                    "our_index":      our_idx,
                    "cust_index":     cust_idx,
                })
                matched_our.add(our_idx)
                matched_cust.add(cust_idx)
                hits += 1

    elapsed = time.time() - t0
    print(f"\n         Done — {hits:,} matches in {elapsed/60:.1f} min")
    return results, matched_our, matched_cust

# ═══════════════════════════════════════════════════════════════════════
# REPORT
# ═══════════════════════════════════════════════════════════════════════

HDR_COLOR       = "1F4E79"
EXACT_COLOR     = "C6EFCE"
FUZZY_COLOR     = "FFEB9C"
UNMATCHED_COLOR = "FCE4D6"

def build_match_rows(matches, our_df, cust_df):
    rows = []
    for m in matches:
        our_row  = our_df.loc[m["our_index"]]
        cust_row = cust_df.loc[m["cust_index"]]
        row = {
            "Match Type":    m.get("Match Type"),
            "Score %":       m.get("Score"),
            "Our Col(s)":    m.get("Our Match Col"),
            "Cust Col(s)":   m.get("Cust Match Col"),
            "Customer File": cust_row.get("_source_file", ""),
            "Customer Sheet":cust_row.get("_source_sheet", ""),
        }
        for k in ["Matched Value", "Category Matched", "Our Text", "Cust Text"]:
            if k in m:
                row[k] = m[k]
        for c in our_df.columns:
            if not c.startswith("_"):
                row[f"OUR_{c}"] = our_row.get(c, "")
        for c in cust_df.columns:
            if not c.startswith("_"):
                row[f"CUST_{c}"] = cust_row.get(c, "")
        rows.append(row)
    return pd.DataFrame(rows)

def format_sheet(ws, header_hex, row_hex=None):
    hdr_fill = PatternFill("solid", fgColor=header_hex)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if row_hex and ws.max_row > 1:
        fill = PatternFill("solid", fgColor=row_hex)
        font = Font(name="Calibri", size=10)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.fill = fill
                cell.font = font
    for col_cells in ws.iter_cols(max_row=min(200, ws.max_row)):
        w = max((len(str(c.value)) for c in col_cells if c.value), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(w + 2, 45)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes    = "A2"

def write_report(exact_df, fuzzy_df, unmatched_our, unmatched_cust,
                 our_df, cust_df, path):
    print("\n📝 Writing report...")
    rows = []
    for fname in cust_df["_source_file"].unique():
        sub     = cust_df[cust_df["_source_file"] == fname]
        n_total = len(sub)
        n_exact = len(exact_df[exact_df["Customer File"] == fname]) if not exact_df.empty else 0
        n_fuzzy = len(fuzzy_df[fuzzy_df["Customer File"] == fname]) if not fuzzy_df.empty else 0
        n_match = n_exact + n_fuzzy
        rows.append({
            "Customer File": fname,
            "Total Rows":    n_total,
            "Exact Matches": n_exact,
            "Fuzzy Matches": n_fuzzy,
            "Total Matched": n_match,
            "Unmatched":     n_total - n_match,
            "Match Rate %":  round(100 * n_match / n_total, 1) if n_total else 0,
        })
    summary_df = pd.DataFrame(rows)
    total = summary_df["Total Rows"].sum()
    summary_df = pd.concat([summary_df, pd.DataFrame([{
        "Customer File": "TOTAL",
        "Total Rows":    total,
        "Exact Matches": summary_df["Exact Matches"].sum(),
        "Fuzzy Matches": summary_df["Fuzzy Matches"].sum(),
        "Total Matched": summary_df["Total Matched"].sum(),
        "Unmatched":     summary_df["Unmatched"].sum(),
        "Match Rate %":  round(100 * summary_df["Total Matched"].sum() / total, 1) if total else 0,
    }])], ignore_index=True)

    drop_our  = [c for c in our_df.columns  if c.startswith("_")]
    drop_cust = [c for c in cust_df.columns if c.startswith("_")]

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        (exact_df.drop(columns=["our_index","cust_index"], errors="ignore")
         if not exact_df.empty
         else pd.DataFrame([{"Info": "No exact matches"}])
         ).to_excel(writer, sheet_name="Exact Matches", index=False)
        (fuzzy_df.drop(columns=["our_index","cust_index"], errors="ignore")
         if not fuzzy_df.empty
         else pd.DataFrame([{"Info": "No fuzzy matches"}])
         ).to_excel(writer, sheet_name="Fuzzy Matches", index=False)
       
    wb = load_workbook(path)
    format_sheet(wb["Summary"],               HDR_COLOR)
    format_sheet(wb["Exact Matches"],         HDR_COLOR, EXACT_COLOR)
    format_sheet(wb["Fuzzy Matches"],         HDR_COLOR, FUZZY_COLOR)
    wb.save(path)
    print(f"✅  Report saved → {path}")

# ═══════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════

def main():
    print("=" * 65)
    print("  MATERIAL MATCHER — Exact + Dynamic Fuzzy")
    print("=" * 65)

    if not os.path.exists(OUR_FILE):
        print(f"\nOur file not found: {OUR_FILE}")
        sys.exit(1)

    print(f"\n📂 Loading our file...")
    our_df = read_file(OUR_FILE)
    print(f"   {len(our_df):,} rows")

    our_exact_ok = [c for c in OUR_EXACT_COLS if c in our_df.columns]
    our_fuzzy_ok = [c for c in OUR_FUZZY_COLS if c in our_df.columns]
    missing_e    = [c for c in OUR_EXACT_COLS if c not in our_df.columns]
    missing_f    = [c for c in OUR_FUZZY_COLS if c not in our_df.columns]

    print(f"\n   YOUR FILE column mapping:")
    print(f"   ┌─ Exact cols  : {our_exact_ok}")
    print(f"   ├─ Fuzzy cols  : {our_fuzzy_ok}")
    print(f"   └─ Cat bucket  : {OUR_CAT_COL}")
    if missing_e: print(f"   ⚠️  Not found (exact): {missing_e}")
    if missing_f: print(f"   ⚠️  Not found (fuzzy): {missing_f}")

    cust_frames = []
    for f in CUSTOMER_FILES:
        if not os.path.exists(f):
            print(f"\n   Missing: {f}")
            continue
        fname = Path(f).name
        print(f"\n📂 Loading: {fname}")
        df = read_file(f)
        if not df.empty:
            print(f"   {len(df):,} rows")
            cust_frames.append(df)

    if not cust_frames:
        print("\nNo customer files loaded.")
        sys.exit(1)

    cust_df = pd.concat(cust_frames, ignore_index=True)
    print(f"\n📊 Total customer rows: {len(cust_df):,}")

    # ── EXACT ──────────────────────────────────────────────────────────
    cust_code_cols = detect_cols(cust_df, CODE_PATTERNS, EXCLUDE_PATTERNS)
    print(f"\n⚡ Exact matching")
    print(f"   Your cols : {our_exact_ok}")
    print(f"   Cust cols : {cust_code_cols}")

    exact_results, m_our_e, m_cust_e = exact_match(
        our_df, our_exact_ok, cust_df, cust_code_cols)
    print(f"   → {len(exact_results):,} exact matches")

    if exact_results:
        col_counts = {}
        for r in exact_results:
            k = f"{r['Our Match Col']} ↔ {r['Cust Match Col']}"
            col_counts[k] = col_counts.get(k, 0) + 1
        for k, v in sorted(col_counts.items(), key=lambda x: -x[1]):
            print(f"      {k}: {v:,}")

    # ── FUZZY — per file ───────────────────────────────────────────────
    print(f"\n🔄 Fuzzy matching (threshold >= {FUZZY_THRESHOLD}%)")
    all_fuzzy = []
    m_our_f, m_cust_f = set(), set()

    for fname in cust_df["_source_file"].unique():
        file_sub    = cust_df[cust_df["_source_file"] == fname]
        n_unmatched = len(file_sub[~file_sub.index.isin(m_cust_e)])
        print(f"\n   📄 {fname}  ({len(file_sub):,} rows, {n_unmatched:,} unmatched)")

        # Dynamic detection for THIS file only
        cust_code_set  = set(detect_cols(file_sub, CODE_PATTERNS, EXCLUDE_PATTERNS))
        cust_name_cols = detect_cols(file_sub, NAME_PATTERNS, EXCLUDE_PATTERNS)
        # Don't use code cols as fuzzy text cols
        cust_name_cols = [c for c in cust_name_cols if c not in cust_code_set]
        cust_cat_col   = pick_cat_col(file_sub)

        print(f"      Exact cols detected : {sorted(cust_code_set) or '(none)'}")
        print(f"      Fuzzy cols detected : {cust_name_cols or '(none)'}")
        print(f"      Cat bucket col      : {cust_cat_col or '(none)'}")

        if not cust_name_cols:
            print(f"      No name/desc columns — skipping fuzzy for this file.")
            continue

        f_res, f_our, f_cust = fuzzy_match_file(
            our_df, our_fuzzy_ok, OUR_CAT_COL,
            file_sub, cust_name_cols, cust_cat_col,
            fname, FUZZY_THRESHOLD,
            skip_our  = m_our_e,
            skip_cust = m_cust_e | m_cust_f,
        )
        print(f"      ✅ {len(f_res):,} fuzzy matches")
        all_fuzzy.extend(f_res)
        m_our_f  |= f_our
        m_cust_f |= f_cust

    # ── Output ─────────────────────────────────────────────────────────
    all_matched_our  = m_our_e | m_our_f
    all_matched_cust = m_cust_e | m_cust_f

    exact_df_out   = build_match_rows(exact_results, our_df, cust_df)
    fuzzy_df_out   = build_match_rows(all_fuzzy,     our_df, cust_df)
    unmatched_our  = our_df.loc[~our_df.index.isin(all_matched_our)]
    unmatched_cust = cust_df.loc[~cust_df.index.isin(all_matched_cust)]

    print(f"\n{'=' * 65}")
    print(f"  FINAL SUMMARY")
    print(f"{'=' * 65}")
    print(f"  Your items:             {len(our_df):,}")
    print(f"  Customer items:         {len(cust_df):,}")
    print(f"  Exact matches:          {len(exact_results):,}")
    print(f"  Fuzzy matches:          {len(all_fuzzy):,}")
    print(f"  Unmatched (ours):       {len(unmatched_our):,}")
    print(f"  Unmatched (customers):  {len(unmatched_cust):,}")

    write_report(exact_df_out, fuzzy_df_out, unmatched_our, unmatched_cust,
                 our_df, cust_df, OUTPUT_FILE)

if __name__ == "__main__":
    main()