"""
Material Matcher — Match your catalog against 6 customer order files.

HOW TO USE:
 1. Place this script in a folder.
 2. Put your file + 6 customer files in the same folder (or edit PATHS below).
 3. pip install openpyxl rapidfuzz pandas xlrd lxml
 4. python material_matcher.py

OUTPUT:  match_report.xlsx  with 4 sheets
  • Exact Matches        – joined on SKU / barcode / material code
  • Fuzzy Matches        – joined on product name / description (score ≥ threshold)
  • Unmatched Ours       – your items with zero matches
  • Unmatched Customers  – customer items with zero matches
"""

import os, re, sys, warnings
from pathlib import Path
from itertools import product as cartesian

import pandas as pd
from rapidfuzz import fuzz, process
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ═══════════════════════════════════════════════════════════════════════
# CONFIGURATION — edit these to match your setup
# ═══════════════════════════════════════════════════════════════════════

# Path to YOUR material / catalog file
OUR_FILE = r"C:\Users\HP\Desktop\functionand sp.csv"

# Paths to the 6 customer order files
CUSTOMER_FILES = [
   r"C:\Users\HP\Downloads\Al Meera July 2024.xlsx",
    r"C:\Users\HP\Downloads\Talabat March.xlsx",
    r"C:\Users\HP\Downloads\fwqniemtselloutdashboardsession2\C4 sep 24.xlsx",
    r"C:\Users\HP\Downloads\fwqniemtselloutdashboardsession2\Grand Mall APR 2024.xlsx",
    r"C:\Users\HP\Downloads\fwqniemtselloutdashboardsession2\QNIE..xlsx",
    r"C:\Users\HP\Downloads\fwqniemtselloutdashboardsession2\RawSparData.xlsx"
]

# Output file
OUTPUT_FILE = r"C:\Users\HP\Desktop\match_report.xlsx"

# Fuzzy match threshold (0–100). 80+ is a strong match.
FUZZY_THRESHOLD = 85

# ═══════════════════════════════════════════════════════════════════════
# COLUMN DETECTION — regex patterns that identify matchable columns
# ═══════════════════════════════════════════════════════════════════════

# Patterns for EXACT-match columns (codes, barcodes, SKUs)
CODE_PATTERNS = [
    r"(?i)\bsku\b",
    r"(?i)\bbarcode\b",
    r"(?i)\bbar[\s_-]?code\b",
    r"(?i)\bupc\b",
    r"(?i)\bean\b",
    r"(?i)\bgtin\b",
    r"(?i)\bitem[\s_-]?(code|no|number|id|#)\b",
    r"(?i)\bproduct[\s_-]?(code|no|number|id|#)\b",
    r"(?i)\bmaterial[\s_-]?(code|no|number|id|#)\b",
    r"(?i)\bpart[\s_-]?(code|no|number|id|#)\b",
    r"(?i)\barticle[\s_-]?(code|no|number|id|#)\b",
    r"(?i)\bcatalog[\s_-]?(code|no|number|id|#)\b",
    r"(?i)\bref(?:erence)?[\s_-]?(code|no|number|id|#)?\b",
    r"(?i)\bhsn\b",
    r"(?i)\bmodel[\s_-]?(code|no|number|id|#)\b",
]

# Patterns for FUZZY-match columns (names, descriptions)
NAME_PATTERNS = [
    r"(?i)\bname\b",
    r"(?i)\bdescription\b",
    r"(?i)\bdesc\b",
    r"(?i)\bproduct\b",
    r"(?i)\bmaterial\b",
    r"(?i)\bitem\b",
    r"(?i)\btitle\b",
    r"(?i)\bgoods\b",
    r"(?i)\bcommodity\b",
]


# ═══════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════

def detect_columns(df, patterns):
    """Return list of column names matching any regex pattern."""
    matched = []
    for col in df.columns:
        col_str = str(col).strip()
        for pat in patterns:
            if re.search(pat, col_str):
                matched.append(col)
                break
    return matched


def normalize(value):
    """Lowercase, strip, collapse whitespace, remove special chars for matching."""
    if pd.isna(value):
        return ""
    s = str(value).strip().lower()
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def detect_engine(filepath):
    """Auto-detect the correct pandas engine based on file content and extension."""
    ext = Path(filepath).suffix.lower()

    # CSV / TSV — not an Excel file at all
    if ext in (".csv", ".tsv"):
        return "csv", ext

    # Try reading the first few bytes to sniff the real format
    with open(filepath, "rb") as f:
        header = f.read(8)

    # XLSX / XLSM / XLSB are ZIP files (start with PK)
    if header[:4] == b"PK\x03\x04":
        if ext == ".xlsb":
            return "pyxlsb", ext
        return "openpyxl", ext

    # Old .xls (OLE2 / Compound Binary) starts with D0 CF 11 E0
    if header[:4] == b"\xd0\xcf\x11\xe0":
        return "xlrd", ext

    # HTML disguised as .xls (common with web-exported "Excel" files)
    # Check first 512 bytes for HTML markers
    with open(filepath, "rb") as f:
        head_512 = f.read(512).lower()
    if b"<html" in head_512 or b"<table" in head_512 or b"<?xml" in head_512:
        return "html", ext

    # Fallback — let pandas guess, but with openpyxl as default
    return "openpyxl", ext


def read_all_sheets(filepath):
    """Read an Excel/CSV file, concatenate all sheets, add source columns."""
    engine, ext = detect_engine(filepath)
    fname = Path(filepath).name
    frames = []

    try:
        # --- CSV / TSV ---
        if engine == "csv":
            sep = "\t" if ext == ".tsv" else ","
            df = pd.read_csv(filepath, dtype=str, sep=sep)
            df.dropna(how="all", inplace=True)
            df.columns = [str(c).strip() for c in df.columns]
            df["_source_file"] = fname
            df["_source_sheet"] = "Sheet1"
            return df

        # --- HTML disguised as Excel ---
        if engine == "html":
            dfs = pd.read_html(filepath, dtype=str)
            for i, df in enumerate(dfs):
                df.dropna(how="all", inplace=True)
                df.columns = [str(c).strip() for c in df.columns]
                df["_source_file"] = fname
                df["_source_sheet"] = f"Table{i+1}"
                frames.append(df)
            if not frames:
                return pd.DataFrame()
            return pd.concat(frames, ignore_index=True)

        # --- Proper Excel (.xlsx, .xls, .xlsm, .xlsb) ---
        xls = pd.ExcelFile(filepath, engine=engine)
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet, dtype=str)
            df.dropna(how="all", inplace=True)
            df.columns = [str(c).strip() for c in df.columns]
            df["_source_file"] = fname
            df["_source_sheet"] = sheet
            frames.append(df)

    except Exception as e:
        print(f"   ⚠️  Error reading {fname}: {e}")
        print(f"       Trying fallback engines...")
        # Last-resort fallback: try every engine
        for fallback in ["openpyxl", "xlrd", "odf"]:
            try:
                xls = pd.ExcelFile(filepath, engine=fallback)
                for sheet in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=sheet, dtype=str)
                    df.dropna(how="all", inplace=True)
                    df.columns = [str(c).strip() for c in df.columns]
                    df["_source_file"] = fname
                    df["_source_sheet"] = sheet
                    frames.append(df)
                print(f"       ✅ Loaded with engine '{fallback}'")
                break
            except Exception:
                continue
        else:
            print(f"   ❌ Could not read {fname} with any engine. Skipping.")
            return pd.DataFrame()

    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


def build_lookup(df, code_cols):
    """Build a dict: normalized_code → list of row indices."""
    lookup = {}
    for idx, row in df.iterrows():
        for col in code_cols:
            val = normalize(row.get(col, ""))
            if val:
                lookup.setdefault(val, []).append(idx)
    return lookup


def exact_match(our_df, our_code_cols, cust_df, cust_code_cols):
    """Find exact matches between our codes and customer codes."""
    results = []
    our_lookup = build_lookup(our_df, our_code_cols)
    matched_our = set()
    matched_cust = set()

    for cust_idx, cust_row in cust_df.iterrows():
        for cust_col in cust_code_cols:
            val = normalize(cust_row.get(cust_col, ""))
            if val and val in our_lookup:
                for our_idx in our_lookup[val]:
                    results.append({
                        "match_type": "EXACT",
                        "matched_on": f"{cust_col} ↔ (code)",
                        "matched_value": val,
                        "score": 100,
                        "our_index": our_idx,
                        "cust_index": cust_idx,
                    })
                    matched_our.add(our_idx)
                    matched_cust.add(cust_idx)
    return results, matched_our, matched_cust


def fuzzy_match(our_df, our_name_cols, cust_df, cust_name_cols,
                threshold, already_matched_our, already_matched_cust):
    """Find fuzzy matches on name/description columns."""
    results = []
    matched_our = set()
    matched_cust = set()

    # Build a combined text for each row in our file
    our_texts = {}
    for idx, row in our_df.iterrows():
        parts = [normalize(row.get(c, "")) for c in our_name_cols]
        combined = " ".join(p for p in parts if p)
        if combined:
            our_texts[idx] = combined

    our_keys = list(our_texts.keys())
    our_vals = list(our_texts.values())

    if not our_vals:
        return results, matched_our, matched_cust

    for cust_idx, cust_row in cust_df.iterrows():
        if cust_idx in already_matched_cust:
            continue
        parts = [normalize(cust_row.get(c, "")) for c in cust_name_cols]
        cust_text = " ".join(p for p in parts if p)
        if not cust_text:
            continue

        # Use rapidfuzz to find top matches
        matches = process.extract(
            cust_text, our_vals,
            scorer=fuzz.token_sort_ratio,
            limit=3,
            score_cutoff=threshold,
        )
        for match_text, score, match_idx_in_list in matches:
            our_idx = our_keys[match_idx_in_list]
            results.append({
                "match_type": "FUZZY",
                "matched_on": "name/description",
                "matched_value": f"'{cust_text}' ≈ '{match_text}'",
                "score": round(score, 1),
                "our_index": our_idx,
                "cust_index": cust_idx,
            })
            matched_our.add(our_idx)
            matched_cust.add(cust_idx)

    return results, matched_our, matched_cust


# ═══════════════════════════════════════════════════════════════════════
# EXCEL REPORT WRITER
# ═══════════════════════════════════════════════════════════════════════

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Arial")
EXACT_FILL = PatternFill("solid", fgColor="C6EFCE")
FUZZY_FILL = PatternFill("solid", fgColor="FCE4D6")
UNMATCHED_FILL = PatternFill("solid", fgColor="F2F2F2")
BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
)
BODY_FONT = Font(name="Arial", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")


def style_sheet(ws, header_fill=HEADER_FILL):
    """Apply formatting to a worksheet."""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = WRAP
            cell.border = BORDER

    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 50), min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 40)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def write_report(exact_rows, fuzzy_rows, unmatched_ours, unmatched_custs, path):
    """Write the 4-sheet Excel report."""
    wb = Workbook()

    # --- Sheet 1: Exact Matches ---
    ws1 = wb.active
    ws1.title = "Exact Matches"
    if exact_rows:
        headers = list(exact_rows[0].keys())
        ws1.append(headers)
        for r in exact_rows:
            ws1.append([r.get(h, "") for h in headers])
        for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
            for cell in row:
                cell.fill = EXACT_FILL if row[0].row % 2 == 0 else PatternFill()
    else:
        ws1.append(["No exact matches found"])
    style_sheet(ws1)

    # --- Sheet 2: Fuzzy Matches ---
    ws2 = wb.create_sheet("Fuzzy Matches")
    if fuzzy_rows:
        headers = list(fuzzy_rows[0].keys())
        ws2.append(headers)
        for r in fuzzy_rows:
            ws2.append([r.get(h, "") for h in headers])
    else:
        ws2.append(["No fuzzy matches found"])
    style_sheet(ws2)

    # --- Sheet 3: Unmatched Ours ---
    ws3 = wb.create_sheet("Unmatched - Ours")
    if not unmatched_ours.empty:
        ws3.append(list(unmatched_ours.columns))
        for _, row in unmatched_ours.iterrows():
            ws3.append([str(v) if pd.notna(v) else "" for v in row])
    else:
        ws3.append(["All our items matched!"])
    style_sheet(ws3)

    # --- Sheet 4: Unmatched Customers ---
    ws4 = wb.create_sheet("Unmatched - Customers")
    if not unmatched_custs.empty:
        ws4.append(list(unmatched_custs.columns))
        for _, row in unmatched_custs.iterrows():
            ws4.append([str(v) if pd.notna(v) else "" for v in row])
    else:
        ws4.append(["All customer items matched!"])
    style_sheet(ws4)

    wb.save(path)
    print(f"\n✅ Report saved → {path}")


# ═══════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  MATERIAL MATCHER — Exact + Fuzzy Cross-File Matching")
    print("=" * 60)

    # --- Load our file ---
    if not os.path.exists(OUR_FILE):
        print(f"\n❌ Our file not found: {OUR_FILE}")
        print("   Edit OUR_FILE at the top of this script.")
        sys.exit(1)

    print(f"\n📂 Loading our file: {OUR_FILE}")
    engine, ext = detect_engine(OUR_FILE)
    print(f"   Format detected: {ext} → engine '{engine}'")
    our_df = read_all_sheets(OUR_FILE)
    print(f"   → {len(our_df)} rows loaded")

    # --- Load customer files ---
    cust_frames = []
    for f in CUSTOMER_FILES:
        if not os.path.exists(f):
            print(f"   ⚠️  Skipping missing file: {f}")
            continue
        engine, ext = detect_engine(f)
        print(f"📂 Loading customer file: {f}  (engine: {engine})")

        df = read_all_sheets(f)
        print(f"   → {len(df)} rows loaded")
        cust_frames.append(df)

    if not cust_frames:
        print("\n❌ No customer files found. Edit CUSTOMER_FILES paths.")
        sys.exit(1)

    cust_df = pd.concat(cust_frames, ignore_index=True)
    print(f"\n📊 Total customer rows: {len(cust_df)}")

    # --- Detect columns ---
    our_code_cols = detect_columns(our_df, CODE_PATTERNS)
    our_name_cols = detect_columns(our_df, NAME_PATTERNS)
    cust_code_cols = detect_columns(cust_df, CODE_PATTERNS)
    cust_name_cols = detect_columns(cust_df, NAME_PATTERNS)

    print(f"\n🔍 Detected columns in OUR file:")
    print(f"   Code columns:  {our_code_cols or '(none found)'}")
    print(f"   Name columns:  {our_name_cols or '(none found)'}")
    print(f"\n🔍 Detected columns in CUSTOMER files:")
    print(f"   Code columns:  {cust_code_cols or '(none found)'}")
    print(f"   Name columns:  {cust_name_cols or '(none found)'}")

    # --- Run exact matching ---
    print(f"\n⚡ Running exact matching...")
    exact_results, matched_our_exact, matched_cust_exact = exact_match(
        our_df, our_code_cols, cust_df, cust_code_cols
    )
    print(f"   → {len(exact_results)} exact matches found")

    # --- Run fuzzy matching ---
    print(f"🔄 Running fuzzy matching (threshold={FUZZY_THRESHOLD})...")
    fuzzy_results, matched_our_fuzzy, matched_cust_fuzzy = fuzzy_match(
        our_df, our_name_cols, cust_df, cust_name_cols,
        FUZZY_THRESHOLD, matched_our_exact, matched_cust_exact,
    )
    print(f"   → {len(fuzzy_results)} fuzzy matches found")

    # --- Build output rows ---
    all_matched_our = matched_our_exact | matched_our_fuzzy
    all_matched_cust = matched_cust_exact | matched_cust_fuzzy

    def build_row(match, our_df, cust_df):
        our_row = our_df.loc[match["our_index"]]
        cust_row = cust_df.loc[match["cust_index"]]
        row = {
            "Match Type": match["match_type"],
            "Score": match["score"],
            "Matched On": match["matched_on"],
            "Matched Value": match["matched_value"],
        }
        for c in our_df.columns:
            if not c.startswith("_"):
                row[f"OUR_{c}"] = our_row.get(c, "")
        for c in cust_df.columns:
            if not c.startswith("_"):
                row[f"CUST_{c}"] = cust_row.get(c, "")
        row["Customer File"] = cust_row.get("_source_file", "")
        row["Customer Sheet"] = cust_row.get("_source_sheet", "")
        return row

    exact_rows = [build_row(m, our_df, cust_df) for m in exact_results]
    fuzzy_rows = [build_row(m, our_df, cust_df) for m in fuzzy_results]

    drop_cols = [c for c in our_df.columns if c.startswith("_")]
    unmatched_ours = our_df.loc[~our_df.index.isin(all_matched_our)].drop(columns=drop_cols, errors="ignore")
    unmatched_custs = cust_df.loc[~cust_df.index.isin(all_matched_cust)].drop(columns=drop_cols, errors="ignore")

    # --- Summary ---
    print(f"\n{'=' * 60}")
    print(f"  SUMMARY")
    print(f"{'=' * 60}")
    print(f"  Our items:              {len(our_df)}")
    print(f"  Customer items:         {len(cust_df)}")
    print(f"  Exact matches:          {len(exact_results)}")
    print(f"  Fuzzy matches:          {len(fuzzy_results)}")
    print(f"  Unmatched (ours):       {len(unmatched_ours)}")
    print(f"  Unmatched (customers):  {len(unmatched_custs)}")

    # --- Write report ---
    write_report(exact_rows, fuzzy_rows, unmatched_ours, unmatched_custs, OUTPUT_FILE)


if __name__ == "__main__":
    main()