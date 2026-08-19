"""
Material Matcher — Exact + Strict Per-Column Fuzzy Matching

ROOT CAUSE FIX:
  Previous version combined ALL detected columns into one giant string.
  "tomato paste" matched "ola tomato paste" vs "luna tomato paste" at 81%
  because both strings contained "tomato paste" — wrong match accepted.

HOW FUZZY NOW WORKS:
  Every column is matched INDIVIDUALLY, never combined into one string.

  Stage 1 — CATEGORY match (one column pair at a time):
    For each our_category_col × each cust_category_col:
      Score = token_sort_ratio(our_val, cust_val)
      Only pairs where score >= CATEGORY_THRESHOLD are kept.
    Best-scoring category column pair wins for each row pair.

  Stage 2 — DESCRIPTION match (on Stage 1 survivors only):
    For each our_description_col × each cust_description_col:
      Score = token_sort_ratio(our_val, cust_val)
      Only pairs where score >= DESCRIPTION_THRESHOLD are kept.
    Best-scoring description column pair wins for each row pair.

  Final result:
    A pair is kept ONLY if it passes BOTH Stage 1 AND Stage 2.
    Final score = min(category_score, description_score)
    — using MIN not average, so a weak link in either stage kills the match.

  If only one stage is possible (file has no category OR no description cols):
    Single stage runs on whichever is available, at SINGLE_THRESHOLD.

THRESHOLDS (tune independently):
  CATEGORY_THRESHOLD    = 85   broader grouping, can afford tighter match
  DESCRIPTION_THRESHOLD = 85   specific item name, must be close
  SINGLE_THRESHOLD      = 88   when only one stage is possible

EXACT MATCHING (all files, unchanged):
  Auto-detect SKU / barcode / code columns.
  Shows exactly which column pair drove each match.

DEDUPLICATION (before any matching):
  3-pass on both your file and every customer file.

HOW TO USE:
  1. Edit OUR_FILE and CUSTOMER_FILES below.
  2. pip install openpyxl rapidfuzz pandas xlrd
  3. python material_matcher.py
"""

import os, re, sys, warnings
from pathlib import Path
import pandas as pd
from rapidfuzz import fuzz, process
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ═══════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════════════

OUR_FILE = r"C:\Users\HP\Desktop\functionand sp.csv"

CUSTOMER_FILES = [
    r"C:\Users\HP\Downloads\Al Meera July 2024.xlsx",
    r"C:\Users\HP\Downloads\Talabat March.xlsx",
    r"C:\Users\HP\Downloads\fwqniemtselloutdashboardsession2\C4 sep 24.xlsx",
    r"C:\Users\HP\Downloads\fwqniemtselloutdashboardsession2\Grand Mall APR 2024.xlsx",
    r"C:\Users\HP\Downloads\fwqniemtselloutdashboardsession2\QNIE..xlsx",
    r"C:\Users\HP\Downloads\fwqniemtselloutdashboardsession2\RawSparData.xlsx",
]

OUTPUT_FILE = r"C:\Users\HP\Desktop\match_report.xlsx"

# ── Thresholds — tune these independently ─────────────────────────────
CATEGORY_THRESHOLD    = 85   # Stage 1: category / group match
DESCRIPTION_THRESHOLD = 85   # Stage 2: item name / description match
SINGLE_THRESHOLD      = 88   # When only one stage possible (no cat or no desc)


# ═══════════════════════════════════════════════════════════════════════
# COLUMN DETECTION PATTERNS
# Columns are detected individually — never combined into one string.
# ═══════════════════════════════════════════════════════════════════════

CODE_PATTERNS = [
    r"(?i)\bsku\b",
    r"(?i)\bbarcode\b",
    r"(?i)\bbar[\s_-]?code\b",
    r"(?i)\bupc\b",
    r"(?i)\bean\b",
    r"(?i)\bgtin\b",
    r"(?i)\bitem[\s_-]?(code|no|number|id|#)\b",
    r"(?i)\bproduct[\s_-]?(code|no|number|id|#)\b",
    r"(?i)\bmaterial[\s_-]?(code|no|number|id|#)\b",
        r"(?i)\bmaterial\b",   

    r"(?i)\bpart[\s_-]?(code|no|number|id|#)\b",
    r"(?i)\barticle[\s_-]?(code|no|number|id|#)\b",
    r"(?i)\bcatalog[\s_-]?(code|no|number|id|#)\b",
    r"(?i)\bref(?:erence)?[\s_-]?(code|no|number|id|#)?\b",
    r"(?i)\bhsn\b",
    r"(?i)\bmodel[\s_-]?(code|no|number|id|#)\b",
]

# Stage 1 — broad product grouping columns
CATEGORY_PATTERNS = [
    r"(?i)\bcategor",        # category, categorylevel, category_name, CATEGORY_NAMNE
    r"(?i)\bfamily\b",       # family, FamilyText
    r"(?i)\bgroup\b",        # product_group, group
    r"(?i)\bdepartment\b",
    r"(?i)\bdept\b",
    r"(?i)\bdivision\b",
    r"(?i)\bsection\b",
    r"(?i)\bmgrp_descr\b",        # keep only this, not generic mgrp
    r"(?i)\bprdh_descr_1\b",
    r"(?i)\bprdh_descr_3\b",
    r"(?i)\bsegment\b",
    r"(?i)\bhierarch",       # hierarchy, prdh_descr_1/2/3
    r"(?i)\bprdh",           # prdh_descr_1, prdh_descr_2
    r"(?i)\bl[1-3]\b",       # L1, L2, L3 hierarchy levels
    r"(?i)\blevel\b",        # Category Level, Analysis Level
    r"(?i)\bclass\b",
    r"(?i)\brange\b",
    r"(?i)\btype\b",
]

# Stage 2 — specific product name / description columns
DESCRIPTION_PATTERNS = [
    r"(?i)\bdescri",
    r"(?i)"                    # description, desc, SU_DESCRIPTION
    r"(?i)\bproduct[\s_-]?name\b",      # Product Name, Product_Name
    r"(?i)\bitem[\s_-]?name\b",         # Item Name
    r"(?i)\bmaterial[\s_-]?desc",       # material_desc
    r"(?i)\bmat[\s_-]?desc",
    r"(?i)^name$",                      # exact "name" column
    r"(?i)\btitle\b",
    r"(?i)\bgoods\b",
    r"(?i)\bcommodity\b",
    r"(?i)\blabel\b",
    r"(?i)\bvariant\b",
    r"(?i)\barticle[\s_-]?desc",
    r"(?i)\bretail[\s_-]?article",      # Retail Article Brand Description Text
]

# Columns that look like descriptions but are actually brand/marketing —
# excluded from DESCRIPTION matching to avoid false positives like
# "Luna" matching "Ola" via a shared category word.
BRAND_EXCLUSION_PATTERNS = [
    r"(?i)^brand$",
    r"(?i)\bbrand[\s_-]?(name|desc|text)\b",
    r"(?i)\bmanufactur",
    r"(?i)\bsupplier\b",
    r"(?i)\bvendor\b",
]


# ═══════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════

def _matches_any(col_name, patterns):
    for pat in patterns:
        if re.search(pat, str(col_name).strip()):
            return True
    return False


def detect_columns(df, patterns):
    return [c for c in df.columns if _matches_any(c, patterns)]


def detect_column_groups(df):
    """
    Assign every column to CODE, CATEGORY, or DESCRIPTION.
    Priority: CODE > CATEGORY > DESCRIPTION.
    Brand-like columns are stripped from DESCRIPTION.
    Returns dict with keys 'code', 'category', 'description'.
    """
    code_cols = detect_columns(df, CODE_PATTERNS)
    code_set  = set(code_cols)

    cat_cols  = [c for c in df.columns
                 if c not in code_set
                 and _matches_any(c, CATEGORY_PATTERNS)]
    cat_set   = set(cat_cols)

    desc_cols = [c for c in df.columns
                 if c not in code_set
                 and c not in cat_set
                 and _matches_any(c, DESCRIPTION_PATTERNS)
                 and not _matches_any(c, BRAND_EXCLUSION_PATTERNS)]

    return {"code": code_cols, "category": cat_cols, "description": desc_cols}


def normalize(value):
    if pd.isna(value):
        return ""
    s = str(value).strip().lower()
    s = re.sub(r"[^\w\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def detect_engine(filepath):
    ext = Path(filepath).suffix.lower()
    if ext in (".csv", ".tsv"):
        return "csv", ext
    with open(filepath, "rb") as f:
        header = f.read(8)
    if header[:4] == b"PK\x03\x04":
        return "pyxlsb" if ext == ".xlsb" else "openpyxl", ext
    if header[:4] == b"\xd0\xcf\x11\xe0":
        return "xlrd", ext
    with open(filepath, "rb") as f:
        head = f.read(512).lower()
    if b"<html" in head or b"<table" in head:
        return "html", ext
    return "openpyxl", ext


# ═══════════════════════════════════════════════════════════════════════
# DEDUPLICATION
# ═══════════════════════════════════════════════════════════════════════

def dedup_dataframe(df, label):
    data_cols      = [c for c in df.columns if not c.startswith("_")]
    removed_frames = []

    # Pass 1 — exact row duplicates
    before = len(df)
    mask   = df.duplicated(subset=data_cols, keep="first")
    r1     = df[mask].copy()
    r1["_removed_reason"] = "Exact duplicate row"
    r1["_removed_from"]   = label
    removed_frames.append(r1)
    df     = df[~mask].reset_index(drop=True)
    if before - len(df):
        print(f"      Pass 1 (exact):       {before - len(df):,} rows removed")

    # Pass 3 — normalised fingerprint
    fps    = df[data_cols].apply(lambda row: "||".join(normalize(v) for v in row), axis=1)
    before = len(df)
    seen, keep = {}, []
    for i, fp in enumerate(fps):
        if fp in seen:
            keep.append(False)
        else:
            seen[fp] = i
            keep.append(True)
    keep_s = pd.Series(keep, dtype=bool)
    r3     = df[~keep_s].copy()
    r3["_removed_reason"] = "Normalised duplicate (whitespace/case/punctuation)"
    r3["_removed_from"]   = label
    removed_frames.append(r3)
    df     = df[keep_s].reset_index(drop=True)
    if before - len(df):
        print(f"      Pass 3 (normalised):  {before - len(df):,} rows removed")

    removed = pd.concat(removed_frames, ignore_index=True) if removed_frames else pd.DataFrame()
    return df, removed


# ═══════════════════════════════════════════════════════════════════════
# FILE READER
# ═══════════════════════════════════════════════════════════════════════

def read_file(filepath):
    engine, ext = detect_engine(filepath)
    fname       = Path(filepath).name
    frames      = []
    try:
        if engine == "csv":
            df = pd.read_csv(filepath, dtype=str,
                             sep="\t" if ext == ".tsv" else ",")
            df.dropna(how="all", inplace=True)
            df.columns          = [str(c).strip() for c in df.columns]
            df["_source_file"]  = fname
            df["_source_sheet"] = "Sheet1"
            return df
        if engine == "html":
            for i, df in enumerate(pd.read_html(filepath, dtype=str)):
                df.dropna(how="all", inplace=True)
                df.columns          = [str(c).strip() for c in df.columns]
                df["_source_file"]  = fname
                df["_source_sheet"] = f"Table{i+1}"
                frames.append(df)
        else:
            xls = pd.ExcelFile(filepath, engine=engine)
            for sheet in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet, dtype=str)
                df.dropna(how="all", inplace=True)
                df.columns          = [str(c).strip() for c in df.columns]
                df["_source_file"]  = fname
                df["_source_sheet"] = sheet
                frames.append(df)
    except Exception as e:
        print(f"   ⚠️  Error reading {fname}: {e} — trying fallbacks...")
        for fallback in ["openpyxl", "xlrd"]:
            try:
                xls = pd.ExcelFile(filepath, engine=fallback)
                for sheet in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=sheet, dtype=str)
                    df.dropna(how="all", inplace=True)
                    df.columns          = [str(c).strip() for c in df.columns]
                    df["_source_file"]  = fname
                    df["_source_sheet"] = sheet
                    frames.append(df)
                print(f"   ✅  Loaded with '{fallback}'")
                break
            except Exception:
                continue
        else:
            print(f"   ❌  Cannot read {fname}. Skipping.")
            return pd.DataFrame()

    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


# ═══════════════════════════════════════════════════════════════════════
# EXACT MATCHING
# ═══════════════════════════════════════════════════════════════════════

def exact_match(our_df, our_code_cols, cust_df, cust_code_cols):
    if not our_code_cols or not cust_code_cols:
        return [], set(), set()

    lookup = {}
    for idx, row in our_df.iterrows():
        for col in our_code_cols:
            val = normalize(row.get(col, ""))
            if val:
                lookup.setdefault(val, []).append((idx, col))

    results, matched_our, matched_cust, seen = [], set(), set(), set()

    for cust_idx, cust_row in cust_df.iterrows():
        for cust_col in cust_code_cols:
            val = normalize(cust_row.get(cust_col, ""))
            if not val or val not in lookup:
                continue
            for our_idx, our_col in lookup[val]:
                pair = (our_idx, cust_idx)
                if pair in seen:
                    continue
                seen.add(pair)
                results.append({
                    "Match Type":     "EXACT",
                    "Our Match Col":  our_col,
                    "Cust Match Col": cust_col,
                    "Matched Value":  val,
                    "Score":          100,
                    "Match Detail":   f"{our_col} ↔ {cust_col} = '{val}'",
                    "our_index":      our_idx,
                    "cust_index":     cust_idx,
                })
                matched_our.add(our_idx)
                matched_cust.add(cust_idx)

    return results, matched_our, matched_cust


# ═══════════════════════════════════════════════════════════════════════
# FUZZY — STRICT PER-COLUMN STAGE
# ═══════════════════════════════════════════════════════════════════════

def _pick_best_col_pair(our_cols, our_sub, cust_cols, cust_sub, sample_n=80):
    """
    Sample-based pre-selection: find the single best (our_col, cust_col)
    pair before doing a full scan.  Tries every combination on up to
    sample_n rows and returns the pair with the highest median score.
    This keeps the full scan O(N) instead of O(cols^2 * N).
    """
    our_valid  = [c for c in our_cols  if c in our_sub.columns]
    cust_valid = [c for c in cust_cols if c in cust_sub.columns]
    if not our_valid or not cust_valid:
        return None, None

    our_s  = our_sub.sample(min(sample_n, len(our_sub)),   random_state=42)
    cust_s = cust_sub.sample(min(sample_n, len(cust_sub)), random_state=42)

    best_pair  = (our_valid[0], cust_valid[0])
    best_score = -1.0

    for oc in our_valid:
        our_vals = [normalize(v) for v in our_s[oc] if normalize(str(v))]
        if not our_vals:
            continue
        for cc in cust_valid:
            scores = []
            for cv in cust_s[cc].dropna():
                cv_n = normalize(cv)
                if not cv_n:
                    continue
                hit = process.extractOne(cv_n, our_vals,
                                         scorer=fuzz.token_sort_ratio)
                if hit:
                    scores.append(hit[1])
            if scores:
                median = sorted(scores)[len(scores) // 2]
                if median > best_score:
                    best_score = median
                    best_pair  = (oc, cc)

    print(f"         Selected pair: '{best_pair[0]}' vs '{best_pair[1]}' (sample median={best_score:.1f})")
    return best_pair


def _best_column_match(our_df, our_cols, cust_df, cust_cols,
                        threshold, stage_label,
                        restrict_our=None, restrict_cust=None):
    """
    Two steps:
      1. _pick_best_col_pair: sample 80 rows to find the best column pair.
      2. Full scan on ONLY that one column pair.

    Never loops cols x cols x rows — always O(N).
    Returns: dict (our_idx, cust_idx) -> {score, our_col, cust_col, our_val, cust_val}
    """
    if not our_cols or not cust_cols:
        return {}

    our_sub  = our_df  if restrict_our  is None else our_df.loc[list(restrict_our)]
    cust_sub = cust_df if restrict_cust is None else cust_df.loc[list(restrict_cust)]

    our_col, cust_col = _pick_best_col_pair(our_cols, our_sub, cust_cols, cust_sub)
    if our_col is None:
        return {}
    if our_col not in our_sub.columns or cust_col not in cust_sub.columns:
        print(f"         [{stage_label}] selected cols missing — skipped")
        return {}

    # Build our lookup
    our_texts = {}
    for idx, row in our_sub.iterrows():
        v = normalize(row.get(our_col, ""))
        if v:
            our_texts[idx] = v

    if not our_texts:
        return {}

    our_keys = list(our_texts.keys())
    our_vals = list(our_texts.values())

    best  = {}
    total = len(cust_sub)

    for i, (cust_idx, cust_row) in enumerate(cust_sub.iterrows()):
        if i % 1000 == 0 and i > 0:
            print(f"         [{stage_label}] {i}/{total}...", end="\r")
        cv = normalize(cust_row.get(cust_col, ""))
        if not cv:
            continue
        hits = process.extract(cv, our_vals,
                               scorer=fuzz.token_sort_ratio,
                               limit=1,
                               score_cutoff=threshold)
        for _, score, pos in hits:
            our_idx = our_keys[pos]
            pair    = (our_idx, cust_idx)
            score_r = round(score, 1)
            if pair not in best or score_r > best[pair]["score"]:
                best[pair] = {
                    "score":    score_r,
                    "our_col":  our_col,
                    "cust_col": cust_col,
                    "our_val":  our_texts[our_idx],
                    "cust_val": cv,
                }

    if total > 1000:
        print()
    return best


# ═══════════════════════════════════════════════════════════════════════
# FUZZY — ORCHESTRATOR
# ═══════════════════════════════════════════════════════════════════════

def fuzzy_match_for_file(our_df, our_groups,
                          cust_file_df,
                          cat_threshold, desc_threshold, single_threshold,
                          skip_our, skip_cust):
    """
    Strict two-stage fuzzy matching for one customer file.

    Both stages use _best_column_match which compares each column
    individually and keeps the best score — never combines columns.

    A pair is kept only if it passes BOTH stages independently.
    Final score = min(cat_score, desc_score) — the weakest link.
    """
    remaining_cust = cust_file_df[~cust_file_df.index.isin(skip_cust)].copy()
    remaining_our  = our_df[~our_df.index.isin(skip_our)].copy()
    fname          = cust_file_df["_source_file"].iloc[0]

    cust_groups = detect_column_groups(cust_file_df)

    our_cat_cols   = our_groups["category"]
    our_desc_cols  = our_groups["description"]
    cust_cat_cols  = cust_groups["category"]
    cust_desc_cols = cust_groups["description"]

    print(f"      Column groups detected:")
    print(f"         YOUR  category cols:    {our_cat_cols  or '(none)'}")
    print(f"         YOUR  description cols: {our_desc_cols or '(none)'}")
    print(f"         CUST  category cols:    {cust_cat_cols  or '(none)'}")
    print(f"         CUST  description cols: {cust_desc_cols or '(none)'}")

    has_cat  = bool(our_cat_cols  and cust_cat_cols)
    has_desc = bool(our_desc_cols and cust_desc_cols)

    results, matched_our, matched_cust, seen = [], set(), set(), set()

    # ── TWO-STAGE: category AND description both available ─────────────
    if has_cat and has_desc:
        print(f"      Mode: STRICT 2-stage  "
              f"[Cat≥{cat_threshold}% AND Desc≥{desc_threshold}%]")

        # Stage 1 — category (per-column, best pair wins)
        print(f"      Stage 1 [Category]  "
              f"{our_cat_cols} vs {cust_cat_cols}...")
        s1 = _best_column_match(remaining_our, our_cat_cols,
                                 remaining_cust, cust_cat_cols,
                                 cat_threshold, "Category")
        print(f"         {len(s1):,} pairs passed category threshold ({cat_threshold}%)")

        if s1:
            # Stage 2 — description, ONLY on Stage 1 survivors
            cand_our  = {p[0] for p in s1}
            cand_cust = {p[1] for p in s1}
            print(f"      Stage 2 [Description]  "
                  f"{our_desc_cols} vs {cust_desc_cols}  "
                  f"(checking {len(cand_cust):,} cust rows)...")
            s2 = _best_column_match(remaining_our, our_desc_cols,
                                     remaining_cust, cust_desc_cols,
                                     desc_threshold, "Description",
                                     restrict_our=cand_our,
                                     restrict_cust=cand_cust)
            print(f"         {len(s2):,} pairs passed description threshold ({desc_threshold}%)")

            # Keep only pairs present in BOTH stages
            common = s1.keys() & s2.keys()
            print(f"         {len(common):,} pairs passed BOTH stages ✅")

            for pair in common:
                if pair in seen:
                    continue
                seen.add(pair)
                p1, p2 = s1[pair], s2[pair]
                final_score = min(p1["score"], p2["score"])
                results.append({
                    "Match Type":        "FUZZY",
                    "Match Stages":      "Category → Description",
                    "Cat Score %":       p1["score"],
                    "Desc Score %":      p2["score"],
                    "Score":             final_score,
                    # Store matched col names as plain strings (one each)
                    "our_cat_col":       p1["our_col"],
                    "our_desc_col":      p2["our_col"],
                    "cust_cat_col":      p1["cust_col"],
                    "cust_desc_col":     p2["cust_col"],
                    "Our Cat Value":     p1["our_val"],
                    "Cust Cat Value":    p1["cust_val"],
                    "Our Desc Value":    p2["our_val"],
                    "Cust Desc Value":   p2["cust_val"],
                    "Match Detail":      (
                        f"[Cat] {p1['our_col']}='{p1['our_val']}' ↔ "
                        f"{p1['cust_col']}='{p1['cust_val']}' ({p1['score']}%)  |  "
                        f"[Desc] {p2['our_col']}='{p2['our_val']}' ↔ "
                        f"{p2['cust_col']}='{p2['cust_val']}' ({p2['score']}%)"
                    ),
                    "our_index":         pair[0],
                    "cust_index":        pair[1],
                })
                matched_our.add(pair[0])
                matched_cust.add(pair[1])

    # ── SINGLE-STAGE: category only ────────────────────────────────────
    elif has_cat:
        print(f"      Mode: 1-stage [Category only, threshold={single_threshold}%]")
        print(f"      {our_cat_cols} vs {cust_cat_cols}...")
        hits = _best_column_match(remaining_our, our_cat_cols,
                                   remaining_cust, cust_cat_cols,
                                   single_threshold, "Category")
        print(f"         {len(hits):,} matches")
        for pair, h in hits.items():
            if pair in seen:
                continue
            seen.add(pair)
            results.append({
                "Match Type":     "FUZZY",
                "Match Stages":   "Category only",
                "Cat Score %":    h["score"],
                "Desc Score %":   "",
                "Score":          h["score"],
                "our_cat_col":    h["our_col"],
                "our_desc_col":   "",
                "cust_cat_col":   h["cust_col"],
                "cust_desc_col":  "",
                "Our Cat Value":  h["our_val"],
                "Cust Cat Value": h["cust_val"],
                "Our Desc Value": "",
                "Cust Desc Value":"",
                "Match Detail":   (
                    f"[Category] {h['our_col']}='{h['our_val']}' ↔ "
                    f"{h['cust_col']}='{h['cust_val']}' ({h['score']}%)"
                ),
                "our_index":      pair[0],
                "cust_index":     pair[1],
            })
            matched_our.add(pair[0])
            matched_cust.add(pair[1])

    # ── SINGLE-STAGE: description only ────────────────────────────────
    elif has_desc:
        print(f"      Mode: 1-stage [Description only, threshold={single_threshold}%]")
        print(f"      {our_desc_cols} vs {cust_desc_cols}...")
        hits = _best_column_match(remaining_our, our_desc_cols,
                                   remaining_cust, cust_desc_cols,
                                   single_threshold, "Description")
        print(f"         {len(hits):,} matches")
        for pair, h in hits.items():
            if pair in seen:
                continue
            seen.add(pair)
            results.append({
                "Match Type":     "FUZZY",
                "Match Stages":   "Description only",
                "Cat Score %":    "",
                "Desc Score %":   h["score"],
                "Score":          h["score"],
                "our_cat_col":    "",
                "our_desc_col":   h["our_col"],
                "cust_cat_col":   "",
                "cust_desc_col":  h["cust_col"],
                "Our Cat Value":  "",
                "Cust Cat Value": "",
                "Our Desc Value": h["our_val"],
                "Cust Desc Value":h["cust_val"],
                "Match Detail":   (
                    f"[Description] {h['our_col']}='{h['our_val']}' ↔ "
                    f"{h['cust_col']}='{h['cust_val']}' ({h['score']}%)"
                ),
                "our_index":      pair[0],
                "cust_index":     pair[1],
            })
            matched_our.add(pair[0])
            matched_cust.add(pair[1])

    else:
        print(f"      ⚠️  No matchable fuzzy columns found — skipping {fname}")

    return results, matched_our, matched_cust


# ═══════════════════════════════════════════════════════════════════════
# SCORE DISTRIBUTION
# ═══════════════════════════════════════════════════════════════════════

def score_distribution(fuzzy_results, threshold):
    lo = int(threshold)
    bands = {
        f"{lo}–{lo+4}%": 0,
        f"{lo+5}–{lo+9}%": 0,
        "90–94%": 0,
        "95–99%": 0,
        "100%": 0,
    }
    for r in fuzzy_results:
        s = r["Score"]
        if   s == 100:     bands["100%"]           += 1
        elif s >= 95:      bands["95–99%"]          += 1
        elif s >= 90:      bands["90–94%"]          += 1
        elif s >= lo + 5:  bands[f"{lo+5}–{lo+9}%"] += 1
        else:              bands[f"{lo}–{lo+4}%"]   += 1
    return bands


# ═══════════════════════════════════════════════════════════════════════
# REPORT
# ═══════════════════════════════════════════════════════════════════════

HDR_COLOR       = "1F4E79"
EXACT_COLOR     = "C6EFCE"
FUZZY_COLOR     = "FFEB9C"
UNMATCHED_COLOR = "FCE4D6"
DUPES_COLOR     = "EDEDED"


def build_match_rows(matches, our_df, cust_df):
    """
    Output ONLY the columns that were actually used for matching.
    No quantity, price, date, or any unrelated column.

    EXACT  → matched code column from your file + matched code column from customer file
    FUZZY  → matched category column(s) + matched description column(s), both sides
    Always included: Match Type, Score, Match Detail, Customer File, Customer Sheet
    """
    rows = []
    for m in matches:
        our_row  = our_df.loc[m["our_index"]]
        cust_row = cust_df.loc[m["cust_index"]]

        row = {
            "Match Type":     m.get("Match Type"),
            "Final Score %":  m.get("Score"),
            "Match Detail":   m.get("Match Detail"),
            "Customer File":  cust_row.get("_source_file", ""),
            "Customer Sheet": cust_row.get("_source_sheet", ""),
        }

        if m.get("Match Type") == "EXACT":
            # Only the two code columns that were identical
            our_col  = m.get("Our Match Col", "")
            cust_col = m.get("Cust Match Col", "")
            row["Matched On"]    = f"{our_col} ↔ {cust_col}"
            row["Matched Value"] = m.get("Matched Value", "")
            if our_col and our_col in our_df.columns:
                row[f"OUR_{our_col}"]   = our_row.get(our_col, "")
            if cust_col and cust_col in cust_df.columns:
                row[f"CUST_{cust_col}"] = cust_row.get(cust_col, "")

        else:
            # Only the category and description columns used in matching
            row["Match Stages"] = m.get("Match Stages", "")
            row["Cat Score %"]  = m.get("Cat Score %", "")
            row["Desc Score %"] = m.get("Desc Score %", "")

            our_cat_col   = m.get("our_cat_col",  "").strip()
            our_desc_col  = m.get("our_desc_col", "").strip()
            cust_cat_col  = m.get("cust_cat_col",  "").strip()
            cust_desc_col = m.get("cust_desc_col", "").strip()

            if our_cat_col and our_cat_col in our_df.columns:
                row[f"OUR_{our_cat_col}"]   = our_row.get(our_cat_col, "")
            if our_desc_col and our_desc_col in our_df.columns:
                row[f"OUR_{our_desc_col}"]  = our_row.get(our_desc_col, "")
            if cust_cat_col and cust_cat_col in cust_df.columns:
                row[f"CUST_{cust_cat_col}"] = cust_row.get(cust_cat_col, "")
            if cust_desc_col and cust_desc_col in cust_df.columns:
                row[f"CUST_{cust_desc_col}"]= cust_row.get(cust_desc_col, "")

        rows.append(row)
    return pd.DataFrame(rows)


def format_sheet(ws, header_hex, row_hex=None):
    hdr_fill  = PatternFill("solid", fgColor=header_hex)
    hdr_font  = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = hdr_align
    if row_hex and ws.max_row > 1:
        fill = PatternFill("solid", fgColor=row_hex)
        font = Font(name="Calibri", size=10)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.fill = fill
                cell.font = font
    for col_cells in ws.iter_cols(max_row=min(200, ws.max_row)):
        w = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(w + 2, 45)
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes    = "A2"


def write_report(exact_df, fuzzy_df, unmatched_our, unmatched_cust,
                 our_df, cust_df, all_removed_df, score_bands,
                 cat_thr, desc_thr, path):
    print("\n📝 Writing report...")

    rows = []
    for fname in cust_df["_source_file"].unique():
        sub       = cust_df[cust_df["_source_file"] == fname]
        n_total   = len(sub)
        n_exact   = len(exact_df[exact_df["Customer File"] == fname]) if not exact_df.empty else 0
        n_fuzzy   = len(fuzzy_df[fuzzy_df["Customer File"] == fname]) if not fuzzy_df.empty else 0
        n_matched = n_exact + n_fuzzy
        n_removed = len(all_removed_df[all_removed_df["_removed_from"] == fname]) \
                    if not all_removed_df.empty else 0
        rows.append({
            "Customer File":      fname,
            "Rows (after dedup)": n_total,
            "Duplicates Removed": n_removed,
            "Exact Matches":      n_exact,
            "Fuzzy Matches":      n_fuzzy,
            "Total Matched":      n_matched,
            "Unmatched":          n_total - n_matched,
            "Match Rate %":       round(100 * n_matched / n_total, 1) if n_total else 0,
        })
    summary_df = pd.DataFrame(rows)
    total_rows = summary_df["Rows (after dedup)"].sum()
    summary_df = pd.concat([summary_df, pd.DataFrame([{
        "Customer File":      "TOTAL",
        "Rows (after dedup)": total_rows,
        "Duplicates Removed": summary_df["Duplicates Removed"].sum(),
        "Exact Matches":      summary_df["Exact Matches"].sum(),
        "Fuzzy Matches":      summary_df["Fuzzy Matches"].sum(),
        "Total Matched":      summary_df["Total Matched"].sum(),
        "Unmatched":          summary_df["Unmatched"].sum(),
        "Match Rate %":       round(100 * summary_df["Total Matched"].sum() / total_rows, 1)
                              if total_rows else 0,
    }])], ignore_index=True)

    # Config note + score bands
    summary_df = pd.concat([summary_df, pd.DataFrame(
        [{"Customer File": f"── Fuzzy Thresholds: Category≥{cat_thr}%  Description≥{desc_thr}% ──"}] +
        [{"Customer File": f"  {b}", "Fuzzy Matches": c} for b, c in score_bands.items()]
    )], ignore_index=True)

    drop_our  = [c for c in our_df.columns  if c.startswith("_")]
    drop_cust = [c for c in cust_df.columns if c.startswith("_")]

    if not all_removed_df.empty:
        mc   = ["_removed_reason", "_removed_from"]
        rest = [c for c in all_removed_df.columns if c not in mc]
        dupes_out = all_removed_df[mc + rest].rename(
            columns={"_removed_reason": "Reason Removed",
                     "_removed_from":   "Removed From"})
    else:
        dupes_out = pd.DataFrame([{"Info": "No duplicates found"}])

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        (exact_df.drop(columns=["our_index","cust_index"], errors="ignore")
         if not exact_df.empty
         else pd.DataFrame([{"Info": "No exact matches found"}])
        ).to_excel(writer, sheet_name="Exact Matches", index=False)
        (fuzzy_df.drop(columns=["our_index","cust_index"], errors="ignore")
         if not fuzzy_df.empty
         else pd.DataFrame([{"Info": "No fuzzy matches found"}])
        ).to_excel(writer, sheet_name="Fuzzy Matches", index=False)
        unmatched_our.drop(columns=drop_our,   errors="ignore").to_excel(
            writer, sheet_name="Unmatched - Ours",      index=False)
        unmatched_cust.drop(columns=drop_cust, errors="ignore").to_excel(
            writer, sheet_name="Unmatched - Customers", index=False)
        dupes_out.to_excel(writer, sheet_name="Duplicates Log", index=False)

    wb = load_workbook(path)
    format_sheet(wb["Summary"],               HDR_COLOR)
    format_sheet(wb["Exact Matches"],         HDR_COLOR, EXACT_COLOR)
    format_sheet(wb["Fuzzy Matches"],         HDR_COLOR, FUZZY_COLOR)
    format_sheet(wb["Unmatched - Ours"],      HDR_COLOR, UNMATCHED_COLOR)
    format_sheet(wb["Unmatched - Customers"], HDR_COLOR, UNMATCHED_COLOR)
    format_sheet(wb["Duplicates Log"],        HDR_COLOR, DUPES_COLOR)
    wb.save(path)
    print(f"✅  Report saved → {path}")


# ═══════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════

def main():
    print("=" * 65)
    print("  MATERIAL MATCHER — Strict Per-Column Fuzzy Matching")
    print("=" * 65)
    print(f"  Thresholds:  Category ≥ {CATEGORY_THRESHOLD}%  |  "
          f"Description ≥ {DESCRIPTION_THRESHOLD}%  |  "
          f"Single-stage ≥ {SINGLE_THRESHOLD}%")

    if not os.path.exists(OUR_FILE):
        print(f"\n❌ Our file not found: {OUR_FILE}")
        sys.exit(1)

    all_removed = []

    # ── Load + dedup OUR file ──────────────────────────────────────────
    print(f"\n📂 Loading our file...")
    our_df = read_file(OUR_FILE)
    print(f"   → {len(our_df):,} rows")
    our_df, our_rem = dedup_dataframe(our_df, "OUR FILE")
    print(f"   ✅ {len(our_df):,} kept  ({len(our_rem):,} duplicates removed)")
    if not our_rem.empty:
        all_removed.append(our_rem)

    our_groups = detect_column_groups(our_df)
    print(f"\n   YOUR file column groups:")
    print(f"      Code cols:        {our_groups['code']        or '(none)'}")
    print(f"      Category cols:    {our_groups['category']    or '(none)'}")
    print(f"      Description cols: {our_groups['description'] or '(none)'}")

    # ── Load + dedup customer files ────────────────────────────────────
    cust_frames = []
    for f in CUSTOMER_FILES:
        if not os.path.exists(f):
            print(f"\n   ⚠️  Missing: {f}")
            continue
        fname = Path(f).name
        print(f"\n📂 Loading: {fname}")
        df = read_file(f)
        if df.empty:
            continue
        print(f"   → {len(df):,} rows")
        df, rem = dedup_dataframe(df, fname)
        print(f"   ✅ {len(df):,} kept  ({len(rem):,} duplicates removed)")
        if not rem.empty:
            all_removed.append(rem)
        cust_frames.append(df)

    if not cust_frames:
        print("\n❌ No customer files loaded.")
        sys.exit(1)

    cust_df = pd.concat(cust_frames, ignore_index=True)
    print(f"\n📊 All customer rows (post-dedup): {len(cust_df):,}")
    print(f"   🔍 Cross-file deduplication...")
    cust_df, cross_rem = dedup_dataframe(cust_df, "CROSS-FILE")
    if not cross_rem.empty:
        print(f"   ✅ {len(cross_rem):,} cross-file duplicates removed")
        all_removed.append(cross_rem)
    else:
        print(f"   ✅ No cross-file duplicates")

    all_removed_df = pd.concat(all_removed, ignore_index=True) \
                     if all_removed else pd.DataFrame()

    # ── EXACT MATCHING ─────────────────────────────────────────────────
    cust_code_cols = detect_columns(cust_df, CODE_PATTERNS)
    print(f"\n⚡ Exact matching:")
    print(f"   Your code cols: {our_groups['code']  or '(none)'}")
    print(f"   Cust code cols: {cust_code_cols or '(none)'}")
    exact_results, m_our_e, m_cust_e = exact_match(
        our_df, our_groups["code"], cust_df, cust_code_cols)
    print(f"   → {len(exact_results):,} exact matches")
    if exact_results:
        col_counts = {}
        for r in exact_results:
            k = f"{r['Our Match Col']} ↔ {r['Cust Match Col']}"
            col_counts[k] = col_counts.get(k, 0) + 1
        for k, v in sorted(col_counts.items(), key=lambda x: -x[1]):
            print(f"      {k}: {v:,}")

    # ── FUZZY MATCHING — per customer file ─────────────────────────────
    print(f"\n🔄 Fuzzy matching — strict per-column, both stages must pass...")
    all_fuzzy   = []
    m_our_f, m_cust_f = set(), set()

    for fname in cust_df["_source_file"].unique():
        file_sub = cust_df[cust_df["_source_file"] == fname]
        print(f"\n   📄 {fname}  ({len(file_sub):,} rows)")

        f_res, f_our, f_cust = fuzzy_match_for_file(
            our_df, our_groups, file_sub,
            CATEGORY_THRESHOLD, DESCRIPTION_THRESHOLD, SINGLE_THRESHOLD,
            skip_our  = m_our_e,
            skip_cust = m_cust_e | m_cust_f,
        )
        print(f"      → {len(f_res):,} fuzzy matches kept")
        all_fuzzy.extend(f_res)
        m_our_f  |= f_our
        m_cust_f |= f_cust

    # ── Score distribution ─────────────────────────────────────────────
    score_bands = score_distribution(all_fuzzy, CATEGORY_THRESHOLD)

    # ── Assemble ───────────────────────────────────────────────────────
    all_matched_our  = m_our_e | m_our_f
    all_matched_cust = m_cust_e | m_cust_f

    exact_df_out   = build_match_rows(exact_results, our_df, cust_df)
    fuzzy_df_out   = build_match_rows(all_fuzzy,     our_df, cust_df)
    unmatched_our  = our_df.loc[~our_df.index.isin(all_matched_our)]
    unmatched_cust = cust_df.loc[~cust_df.index.isin(all_matched_cust)]

    # ── Console summary ────────────────────────────────────────────────
    print(f"\n{'=' * 65}")
    print(f"  FINAL SUMMARY")
    print(f"{'=' * 65}")
    print(f"  Your items (after dedup):       {len(our_df):,}")
    print(f"  Customer items (after dedup):   {len(cust_df):,}")
    print(f"  Total duplicates removed:       {len(all_removed_df):,}")
    print(f"  Exact matches:                  {len(exact_results):,}")
    print(f"  Fuzzy matches:                  {len(all_fuzzy):,}")
    print(f"  Unmatched (ours):               {len(unmatched_our):,}")
    print(f"  Unmatched (customers):          {len(unmatched_cust):,}")
    print(f"\n  Fuzzy Score Bands:")
    for band, count in score_bands.items():
        bar = "█" * min(count, 50)
        print(f"    {band:>8}  {bar}  {count:,}")

    write_report(exact_df_out, fuzzy_df_out, unmatched_our, unmatched_cust,
                 our_df, cust_df, all_removed_df, score_bands,
                 CATEGORY_THRESHOLD, DESCRIPTION_THRESHOLD, OUTPUT_FILE)


if __name__ == "__main__":
    main()