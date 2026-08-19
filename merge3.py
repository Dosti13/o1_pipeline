"""
Material Matcher — Match your catalog against 6 customer order files.

FIX: Report writing now uses pandas ExcelWriter (fast bulk write)
     instead of cell-by-cell openpyxl (was hanging on large datasets).

HOW TO USE:
  1. Edit OUR_FILE and CUSTOMER_FILES paths below.
  2. pip install openpyxl rapidfuzz pandas xlrd
  3. python material_matcher.py

OUTPUT: match_report.xlsx with 5 sheets
  • Summary              – counts per customer file
  • Exact Matches        – joined on SKU / barcode / material code
  • Fuzzy Matches        – joined on product name / description
  • Unmatched Ours       – your items with zero matches
  • Unmatched Customers  – customer items with zero matches
"""

import os, re, sys, warnings
from pathlib import Path
import pandas as pd
from rapidfuzz import fuzz, process
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ═══════════════════════════════════════════════════════════════════════
# CONFIGURATION — edit these paths
# ═══════════════════════════════════════════════════════════════════════

OUR_FILE = r"C:\Users\HP\Desktop\functionand sp.csv"

CUSTOMER_FILES = [
    r"C:\Users\HP\Downloads\Al Meera July 2024.xlsx",
    r"C:\Users\HP\Downloads\Talabat March.xlsx",
    r"C:\Users\HP\Downloads\fwqniemtselloutdashboardsession2\C4 sep 24.xlsx",
    r"C:\Users\HP\Downloads\fwqniemtselloutdashboardsession2\Grand Mall APR 2024.xlsx",
    r"C:\Users\HP\Downloads\fwqniemtselloutdashboardsession2\QNIE..xlsx",
    r"C:\Users\HP\Downloads\fwqniemtselloutdashboardsession2\RawSparData.xlsx",
]

OUTPUT_FILE = r"C:\Users\HP\Desktop\match_report.xlsx"

FUZZY_THRESHOLD = 80   # 0–100. Lower = more matches but less accurate.

# ── OVERRIDE: pin exact column names from YOUR file for fuzzy matching ─
# These are used instead of auto-detection on the OUR side.
# Add or remove columns as needed.
OUR_NAME_COLS_OVERRIDE = [
    "material_desc",   # sub-category  ← your actual column
    "mgrp_descr",      # category       ← your actual column
]

# ═══════════════════════════════════════════════════════════════════════
# COLUMN DETECTION
# ═══════════════════════════════════════════════════════════════════════

CODE_PATTERNS = [
    r"(?i)\bsku\b", r"(?i)\bbarcode\b", r"(?i)\bbar[\s_-]?code\b",
    r"(?i)\bupc\b", r"(?i)\bean\b", r"(?i)\bgtin\b",
    r"(?i)\bitem[\s_-]?(code|no|number|id|#)\b",
    r"(?i)\bproduct[\s_-]?(code|no|number|id|#)\b",
    r"(?i)\bmaterial[\s_-]?(code|no|number|id|#)\b",
    r"(?i)\bpart[\s_-]?(code|no|number|id|#)\b",
    r"(?i)\barticle[\s_-]?(code|no|number|id|#)\b",
    r"(?i)\bcatalog[\s_-]?(code|no|number|id|#)\b",
    r"(?i)\bref(?:erence)?[\s_-]?(code|no|number|id|#)?\b",
    r"(?i)\bhsn\b", r"(?i)\bmodel[\s_-]?(code|no|number|id|#)\b",
]

NAME_PATTERNS = [
    r"(?i)\bname\b", r"(?i)\bdescription\b", r"(?i)desc",
    r"(?i)\bproduct\b", r"(?i)material", r"(?i)\bitem\b",
    r"(?i)\btitle\b", r"(?i)\bgoods\b", r"(?i)\bcommodity\b",
    r"(?i)mgrp", r"(?i)brand", r"(?i)categ",
]


# ═══════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════

def detect_columns(df, patterns):
    matched = []
    for col in df.columns:
        col_str = str(col).strip()
        for pat in patterns:
            if re.search(pat, col_str):
                matched.append(col)
                break
    return matched


def normalize(value):
    if pd.isna(value):
        return ""
    s = str(value).strip().lower()
    s = re.sub(r"[^\w\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def detect_engine(filepath):
    ext = Path(filepath).suffix.lower()
    if ext in (".csv", ".tsv"):
        return "csv", ext
    with open(filepath, "rb") as f:
        header = f.read(8)
    if header[:4] == b"PK\x03\x04":
        return "pyxlsb" if ext == ".xlsb" else "openpyxl", ext
    if header[:4] == b"\xd0\xcf\x11\xe0":
        return "xlrd", ext
    with open(filepath, "rb") as f:
        head = f.read(512).lower()
    if b"<html" in head or b"<table" in head:
        return "html", ext
    return "openpyxl", ext


def read_file(filepath):
    engine, ext = detect_engine(filepath)
    fname = Path(filepath).name
    frames = []
    try:
        if engine == "csv":
            sep = "\t" if ext == ".tsv" else ","
            df = pd.read_csv(filepath, dtype=str, sep=sep)
            df.dropna(how="all", inplace=True)
            df.columns = [str(c).strip() for c in df.columns]
            df["_source_file"] = fname
            df["_source_sheet"] = "Sheet1"
            return df
        if engine == "html":
            dfs = pd.read_html(filepath, dtype=str)
            for i, df in enumerate(dfs):
                df.dropna(how="all", inplace=True)
                df.columns = [str(c).strip() for c in df.columns]
                df["_source_file"] = fname
                df["_source_sheet"] = f"Table{i+1}"
                frames.append(df)
        else:
            xls = pd.ExcelFile(filepath, engine=engine)
            for sheet in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet, dtype=str)
                df.dropna(how="all", inplace=True)
                df.columns = [str(c).strip() for c in df.columns]
                df["_source_file"] = fname
                df["_source_sheet"] = sheet
                frames.append(df)
    except Exception as e:
        print(f"   ⚠️  Error reading {fname}: {e} — trying fallbacks...")
        for fallback in ["openpyxl", "xlrd"]:
            try:
                xls = pd.ExcelFile(filepath, engine=fallback)
                for sheet in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=sheet, dtype=str)
                    df.dropna(how="all", inplace=True)
                    df.columns = [str(c).strip() for c in df.columns]
                    df["_source_file"] = fname
                    df["_source_sheet"] = sheet
                    frames.append(df)
                print(f"   ✅  Loaded with '{fallback}'")
                break
            except Exception:
                continue
        else:
            print(f"   ❌  Cannot read {fname}. Skipping.")
            return pd.DataFrame()
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


# ═══════════════════════════════════════════════════════════════════════
# MATCHING
# ═══════════════════════════════════════════════════════════════════════

def exact_match(our_df, our_code_cols, cust_df, cust_code_cols):
    if not our_code_cols or not cust_code_cols:
        return [], set(), set()

    lookup = {}
    for idx, row in our_df.iterrows():
        for col in our_code_cols:
            val = normalize(row.get(col, ""))
            if val:
                lookup.setdefault(val, []).append(idx)

    results, matched_our, matched_cust = [], set(), set()
    for cust_idx, cust_row in cust_df.iterrows():
        for cust_col in cust_code_cols:
            val = normalize(cust_row.get(cust_col, ""))
            if val and val in lookup:
                for our_idx in lookup[val]:
                    results.append({
                        "Match Type": "EXACT",
                        "Matched On": cust_col,
                        "Matched Value": val,
                        "Score": 100,
                        "our_index": our_idx,
                        "cust_index": cust_idx,
                    })
                    matched_our.add(our_idx)
                    matched_cust.add(cust_idx)
    return results, matched_our, matched_cust


def fuzzy_match(our_df, our_name_cols, cust_df, cust_name_cols,
                threshold, skip_our, skip_cust):
    if not our_name_cols or not cust_name_cols:
        return [], set(), set()

    our_texts = {}
    for idx, row in our_df.iterrows():
        parts = [normalize(row.get(c, "")) for c in our_name_cols]
        combined = " ".join(p for p in parts if p)
        if combined:
            our_texts[idx] = combined

    our_keys = list(our_texts.keys())
    our_vals = list(our_texts.values())

    results, matched_our, matched_cust = [], set(), set()
    total = len(cust_df)
    for i, (cust_idx, cust_row) in enumerate(cust_df.iterrows()):
        if i % 500 == 0:
            print(f"   Fuzzy: {i}/{total} rows processed...", end="\r")
        if cust_idx in skip_cust:
            continue
        parts = [normalize(cust_row.get(c, "")) for c in cust_name_cols]
        cust_text = " ".join(p for p in parts if p)
        if not cust_text:
            continue

        matches = process.extract(
            cust_text, our_vals,
            scorer=fuzz.token_sort_ratio,
            limit=3,
            score_cutoff=threshold,
        )
        for match_text, score, match_pos in matches:
            our_idx = our_keys[match_pos]
            results.append({
                "Match Type": "FUZZY",
                "Matched On": "name/description",
                "Matched Value": f"{cust_text} ≈ {match_text}",
                "Score": round(score, 1),
                "our_index": our_idx,
                "cust_index": cust_idx,
            })
            matched_our.add(our_idx)
            matched_cust.add(cust_idx)
    print()
    return results, matched_our, matched_cust


# ═══════════════════════════════════════════════════════════════════════
# REPORT — uses pandas bulk write (FAST) + light openpyxl formatting
# ═══════════════════════════════════════════════════════════════════════

HDR_COLOR = "1F4E79"
EXACT_COLOR = "C6EFCE"
FUZZY_COLOR = "FFEB9C"
UNMATCHED_COLOR = "FCE4D6"


def build_match_rows(matches, our_df, cust_df):
    rows = []
    for m in matches:
        our_row = our_df.loc[m["our_index"]]
        cust_row = cust_df.loc[m["cust_index"]]
        row = {
            "Match Type": m["Match Type"],
            "Score": m["Score"],
            "Matched On": m["Matched On"],
            "Matched Value": m["Matched Value"],
            "Customer File": cust_row.get("_source_file", ""),
            "Customer Sheet": cust_row.get("_source_sheet", ""),
        }
        for c in our_df.columns:
            if not c.startswith("_"):
                row[f"OUR_{c}"] = our_row.get(c, "")
        for c in cust_df.columns:
            if not c.startswith("_"):
                row[f"CUST_{c}"] = cust_row.get(c, "")
        rows.append(row)
    return pd.DataFrame(rows)


def format_sheet(ws, header_hex, row_hex=None):
    """Apply header style and optional row fill — column-by-column for speed."""
    hdr_fill = PatternFill("solid", fgColor=header_hex)
    hdr_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align

    if row_hex and ws.max_row > 1:
        row_fill = PatternFill("solid", fgColor=row_hex)
        body_font = Font(name="Arial", size=10)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.fill = row_fill
                cell.font = body_font

    # Auto-width (sample first 200 rows only — keeps it fast)
    for col_cells in ws.iter_cols(max_row=min(200, ws.max_row)):
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=8,
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 40)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def write_report(exact_df, fuzzy_df, unmatched_our, unmatched_cust,
                 our_df, cust_df, path):
    print("\n📝 Writing report (this may take a moment for large files)...")

    # --- Summary per customer file ---
    summary_rows = []
    for fname in cust_df["_source_file"].unique():
        cust_sub = cust_df[cust_df["_source_file"] == fname]
        n_total = len(cust_sub)
        n_exact = len(exact_df[exact_df["Customer File"] == fname]) if not exact_df.empty else 0
        n_fuzzy = len(fuzzy_df[fuzzy_df["Customer File"] == fname]) if not fuzzy_df.empty else 0
        n_matched = n_exact + n_fuzzy
        summary_rows.append({
            "Customer File": fname,
            "Total Customer Rows": n_total,
            "Exact Matches": n_exact,
            "Fuzzy Matches": n_fuzzy,
            "Total Matched": n_matched,
            "Unmatched": n_total - n_matched,
            "Match Rate %": round(100 * n_matched / n_total, 1) if n_total else 0,
        })
    summary_df = pd.DataFrame(summary_rows)

    # Totals row
    totals = {
        "Customer File": "TOTAL",
        "Total Customer Rows": summary_df["Total Customer Rows"].sum(),
        "Exact Matches": summary_df["Exact Matches"].sum(),
        "Fuzzy Matches": summary_df["Fuzzy Matches"].sum(),
        "Total Matched": summary_df["Total Matched"].sum(),
        "Unmatched": summary_df["Unmatched"].sum(),
        "Match Rate %": round(100 * summary_df["Total Matched"].sum() / summary_df["Total Customer Rows"].sum(), 1),
    }
    summary_df = pd.concat([summary_df, pd.DataFrame([totals])], ignore_index=True)

    drop_cols = [c for c in our_df.columns if c.startswith("_")]
    unmatched_our_out = unmatched_our.drop(columns=drop_cols, errors="ignore")
    drop_cols2 = [c for c in cust_df.columns if c.startswith("_")]
    unmatched_cust_out = unmatched_cust.drop(columns=drop_cols2, errors="ignore")

    # Bulk write all sheets via pandas
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        if not exact_df.empty:
            exact_df.drop(columns=["our_index", "cust_index"], errors="ignore").to_excel(
                writer, sheet_name="Exact Matches", index=False)
        else:
            pd.DataFrame([{"Info": "No exact matches found"}]).to_excel(
                writer, sheet_name="Exact Matches", index=False)
        if not fuzzy_df.empty:
            fuzzy_df.drop(columns=["our_index", "cust_index"], errors="ignore").to_excel(
                writer, sheet_name="Fuzzy Matches", index=False)
        else:
            pd.DataFrame([{"Info": "No fuzzy matches found"}]).to_excel(
                writer, sheet_name="Fuzzy Matches", index=False)
        unmatched_our_out.to_excel(writer, sheet_name="Unmatched - Ours", index=False)
        unmatched_cust_out.to_excel(writer, sheet_name="Unmatched - Customers", index=False)

    # Light formatting pass
    wb = load_workbook(path)
    format_sheet(wb["Summary"],            HDR_COLOR)
    format_sheet(wb["Exact Matches"],      HDR_COLOR, EXACT_COLOR)
    format_sheet(wb["Fuzzy Matches"],      HDR_COLOR, FUZZY_COLOR)
    format_sheet(wb["Unmatched - Ours"],   HDR_COLOR, UNMATCHED_COLOR)
    format_sheet(wb["Unmatched - Customers"], HDR_COLOR, UNMATCHED_COLOR)
    wb.save(path)
    print(f"✅ Report saved → {path}")


# ═══════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  MATERIAL MATCHER — Exact + Fuzzy Cross-File Matching")
    print("=" * 60)

    if not os.path.exists(OUR_FILE):
        print(f"\n❌ Our file not found: {OUR_FILE}")
        sys.exit(1)

    print(f"\n📂 Loading our file...")
    our_df = read_file(OUR_FILE)
    print(f"   → {len(our_df):,} rows loaded")

    cust_frames = []
    for f in CUSTOMER_FILES:
        if not os.path.exists(f):
            print(f"   ⚠️  Missing: {f}")
            continue
        print(f"📂 Loading: {Path(f).name}")
        df = read_file(f)
        if not df.empty:
            print(f"   → {len(df):,} rows")
            cust_frames.append(df)

    if not cust_frames:
        print("\n❌ No customer files loaded.")
        sys.exit(1)

    cust_df = pd.concat(cust_frames, ignore_index=True)
    print(f"\n📊 Total customer rows: {len(cust_df):,}")

    our_code_cols  = detect_columns(our_df, CODE_PATTERNS)
    cust_code_cols = detect_columns(cust_df, CODE_PATTERNS)
    cust_name_cols = detect_columns(cust_df, NAME_PATTERNS)

    # Use override for our name cols if set, else auto-detect
    if OUR_NAME_COLS_OVERRIDE:
        our_name_cols = [c for c in OUR_NAME_COLS_OVERRIDE if c in our_df.columns]
        missing = [c for c in OUR_NAME_COLS_OVERRIDE if c not in our_df.columns]
        if missing:
            print(f"   ⚠️  OUR_NAME_COLS_OVERRIDE — not found in file: {missing}")
    else:
        our_name_cols = detect_columns(our_df, NAME_PATTERNS)

    print(f"\n🔍 Our file   — code cols: {our_code_cols or '(none)'}")
    print(f"              — name cols: {our_name_cols or '(none)'}")
    print(f"🔍 Cust files — code cols: {cust_code_cols or '(none)'}")
    print(f"              — name cols: {cust_name_cols or '(none)'}")

    print(f"\n⚡ Running exact matching...")
    exact_results, m_our_e, m_cust_e = exact_match(
        our_df, our_code_cols, cust_df, cust_code_cols)
    print(f"   → {len(exact_results):,} exact matches")

    print(f"🔄 Running fuzzy matching (threshold={FUZZY_THRESHOLD})...")
    fuzzy_results, m_our_f, m_cust_f = fuzzy_match(
        our_df, our_name_cols, cust_df, cust_name_cols,
        FUZZY_THRESHOLD, m_our_e, m_cust_e)
    print(f"   → {len(fuzzy_results):,} fuzzy matches")

    all_matched_our  = m_our_e  | m_our_f
    all_matched_cust = m_cust_e | m_cust_f

    exact_df = build_match_rows(exact_results, our_df, cust_df)
    fuzzy_df = build_match_rows(fuzzy_results, our_df, cust_df)
    unmatched_our  = our_df.loc[~our_df.index.isin(all_matched_our)]
    unmatched_cust = cust_df.loc[~cust_df.index.isin(all_matched_cust)]

    print(f"\n{'=' * 60}")
    print(f"  SUMMARY")
    print(f"{'=' * 60}")
    print(f"  Our items:              {len(our_df):,}")
    print(f"  Customer items:         {len(cust_df):,}")
    print(f"  Exact matches:          {len(exact_results):,}")
    print(f"  Fuzzy matches:          {len(fuzzy_results):,}")
    print(f"  Unmatched (ours):       {len(unmatched_our):,}")
    print(f"  Unmatched (customers):  {len(unmatched_cust):,}")

    write_report(exact_df, fuzzy_df, unmatched_our, unmatched_cust,
                 our_df, cust_df, OUTPUT_FILE)


if __name__ == "__main__":
    main()