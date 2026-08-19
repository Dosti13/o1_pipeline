"""
resolve_brand_and_category.py
==============================
Rule-based (no ML / no LLM calls) clean-up for dim_material_master_qnie_csv.xlsx.

    python resolve_brand_and_category.py <input_xlsx> [output_xlsx]

WHAT THIS DOES
--------------

1) BRAND - fills in `brand_standardized == UNKNOWN` rows using a 4-step
   cascade, in priority order. Every filled row records *how* it was
   filled (`brand_resolution_method`) and *why* (`brand_resolution_evidence`)
   so every single change is auditable and reversible.

     Step A1 - mgrp_descr SIBLING CONSENSUS (strongest signal)
        Other rows that already carry a confirmed brand and share the exact
        same material-group description (mgrp_descr) get a reliability-
        weighted vote (barcode > rule_exact > rule_fuzzy > llm). This is
        exactly your own ZWAN example: mgrp_descr 'ZWAN' already has plenty
        of barcode/rule_exact-confirmed siblings tagged brand=ZWAN, so an
        UNKNOWN row sharing that same mgrp_descr inherits ZWAN too. It also
        self-corrects noise already in the file - e.g. one mgrp_descr group
        had some 'llm'-tagged rows guessing brand 'TIFF' while its far more
        numerous rule_exact/barcode-tagged siblings said 'TIFFANY' - the
        higher-weighted, larger-consensus answer (TIFFANY) wins.

     Step A2 - mgrp_descr TEXT MATCH (fallback when there's no sibling
        consensus, e.g. the row is the only one of its kind)
        Matches the mgrp_descr text itself (and its punctuation-separated
        segments, so 'BK INGR-PURATOS' -> PURATOS) against a dictionary of
        every already-confirmed brand in the file.

     Step B - material_desc LEADING WORDS
        Skips logistics/state/origin prefixes (FZ, CH, USA, AUS, JOR, KW...)
        then checks the next 1-4 words against the same brand dictionary.

     Step C - material_desc FULL SCAN
        Scans the rest of the description for any recognized brand name.
        Longest match wins; if the scan turns up more than one genuinely
        different brand, the row is flagged 'ambiguous' rather than guessed.

   A GENERIC-TERM BLACKLIST (curated + auto-derived from your own 135
   category names) stops category/description words - FRESH, BEEF, CHICKEN,
   CH, FZ, USA, TILAPIA, MOZZARELLA, ROYAL, SUPER, DOHA, ... - from ever
   being mistaken for a brand. This list is what your own example was
   getting at ("fresh and beef... are not brand"): mgrp_descr 'ZWAN' passes
   because ZWAN isn't generic wording, but mgrp_descr 'FROZEN CHICKEN' never
   will, because every word in it is on the blacklist.

   A MINIMUM-SUPPORT FILTER also drops one-off, low-trust guesses (e.g. a
   single 'llm' mis-tag like 'CH PF' or 'NET') from the trusted dictionary
   unless they are barcode-verified - a real brand almost always shows up on
   more than a couple of SKUs; a stray mis-tag usually doesn't.

2) CATEGORY - audits `lulu_category`, but only touches rows that aren't
   already trustworthy:
     - barcode-matched rows and high-confidence (>=0.80) rows are left
       completely untouched.
     - non-food rows (cleaning supplies, coffee-machine parts, disposables,
       toiletries...) are recognized as such and marked "not applicable" -
       your 135-category list is a FOOD taxonomy, so UNKNOWN is the
       *correct* answer for these, not something to force-fit.
     - everything else gets scored against keyword profiles built ONLY from
       the rows we already trust, using specificity-weighted word overlap
       (a word that shows up under many categories counts for less than one
       that is unique to a single category). A row is only ever reassigned
       to one of the EXISTING 135 categories, never a new one, and only
       when there's a clear, unambiguous winner. Anything else is left as
       NEEDS_MANUAL_REVIEW rather than guessed.

Everything above is deterministic string / keyword logic - there are no
external API calls and nothing here depends on network access, so the
script's output is fully reproducible.

OUTPUT WORKBOOK
---------------
  standardized_master   original data + 6 new columns (see below), with
                         brand_final==UNKNOWN highlighted red and any
                         category flag needing attention highlighted
                         amber/green
  brand_crosswalk        unchanged, passed through from the input
  needs_manual_review     just the rows still needing a human look
  summary                 methodology notes + before/after counts

New columns added to standardized_master:
  brand_final                  the resolved brand (existing value kept as-is
                                if it wasn't UNKNOWN)
  brand_resolution_method      existing / mgrp_descr_sibling_consensus /
                                mgrp_descr_text_match / desc_prefix_match /
                                desc_scan_match / ambiguous_multiple_brands /
                                unresolved
  brand_resolution_evidence    human-readable reason for the assignment
  lulu_category_final          the resolved category (existing value kept
                                as-is unless reassigned)
  category_audit_flag          TRUSTED_BARCODE / OK_HIGH_CONFIDENCE /
                                OK_MEDIUM_CONFIDENCE / OK_LOW_CONFIDENCE_CONFIRMED /
                                NOT_FOOD_NO_CATEGORY_APPLICABLE /
                                REASSIGNED_NEAREST_MATCH /
                                NEEDS_MANUAL_REVIEW_NO_MATCH /
                                NEEDS_MANUAL_REVIEW_AMBIGUOUS
  category_audit_evidence      human-readable reason for the flag

TUNING
------
All thresholds live near the top of the file (HIGH_CONF, MED_CONF, MIN_SCORE,
CLEAR_WINNER_RATIO, min_support_non_barcode) and the word lists
(CURATED_GENERIC, COUNTRY_TERMS, SEAFOOD_TERMS, FLAVOUR_TERMS,
CHEESE_TYPE_TERMS, MARKETING_TERMS, PLACE_TERMS, NON_FOOD_MGRP_SIGNALS) are
plain space-separated strings - add a word and re-run to tighten or loosen
any rule without touching the logic itself.
"""
import re
import sys
from collections import Counter, defaultdict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ==========================================================================
# TUNABLE THRESHOLDS
# ==========================================================================
HIGH_CONF = 0.80              # >= this -> category considered trustworthy
MED_CONF = 0.65                # >= this -> "OK but worth a spot check"
CLEAR_WINNER_RATIO = 1.8       # best category score must beat runner-up by this multiple
MIN_SCORE = 2.0                 # minimum absolute keyword-overlap score to trust a category suggestion
MIN_SUPPORT_NON_BARCODE = 3     # a non-barcode-verified brand needs >= this many occurrences to be trusted globally
CONSENSUS_SHARE = 0.6           # a sibling-consensus brand needs >= this share of the weighted vote

# ==========================================================================
# 1. TEXT HELPERS
# ==========================================================================
WORD_RE = re.compile(r"[A-Z0-9]+")
UNIT_TOKEN_RE = re.compile(r"^\d+([A-Z]{0,4})?$")


def tight(s):
    """Uppercase + strip every non-alphanumeric char, so 'AL WATANIA' and
    'ALWATANIA' collapse to the same lookup key."""
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    return re.sub(r"[^A-Z0-9]", "", str(s).upper())


def words(s):
    """Uppercase word tokens."""
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return []
    return WORD_RE.findall(str(s).upper())


def ngrams(tok_list, max_n=4):
    """Yield (start_idx, length, ngram_words) for n = max_n down to 1,
    longest first, so a longer brand match wins over a shorter partial one."""
    n_tok = len(tok_list)
    for n in range(min(max_n, n_tok), 0, -1):
        for i in range(0, n_tok - n + 1):
            yield i, n, tok_list[i:i + n]


# ==========================================================================
# 2. GENERIC / NON-BRAND VOCABULARY (curated + auto-derived from the data)
# ==========================================================================
CURATED_GENERIC = """
FRESH FROZEN FZ CH CHILLED CANNED DRIED ORGANIC NATURAL PURE LOCAL IMPORTED IMP
PREMIUM CLASSIC ORIGINAL MIXED MIX MIXES ASSORTED WHOLE WHL SLICED MINCED GROUND
DICED CUBED BONELESS BNLS SKINLESS SKLS BREADED MARINATED PROCESS PROCESSED
ITEMS ITEM PRODUCT PRODUCTS RANGE OTHER OTHERS OTHR NPD COMMN COMMON LOW VALUE
HIGH BY AIR SEA NON ACTIVE BAKERY BAKING INGR INGREDIENT INGREDIENTS BK FSD
QNIE SF BRZ IND AUS USA UAE UK NZ KSA GCC JOR KW ARG PLND POL TUR CHINA RUS
OMN QAR THAI KOREAN KOREA PHILIPPINE PHILLPINO S AFRICA IH BF BL SL RTE RTC
IQF CV CTN PCS PKT GM GR KG ML LTR L X CO CO. LTD FZN
MEAT POULTRY FISH SEAFOOD VEGETABLE VEGETABLES VEG FRUIT FRUITS HERBS SPICE
SPICES SEASONING SAUCE SAUCES PASTE DAIRY DIARY BUTCHER MY DETERGENT CLEANING
DISPOSABLE MACHINE MACHN MCHN ACCESSORIES ACCS TOOLS EQUIPMENT BARISTA HOT COLD
DRINK DRINKS BEVERAGE BEVERAGES JUICE JUICES WATER TEA COFFEE WAGYU ANGUS
BEEF LAMB MUTTON VEAL CHICKEN CKN CHK TURKEY DUCK QUAIL QUAILS PRAWN PRAWNS
SHRIMP SHRIMPS EGG EGGS CHEESE MILK YOGURT YOGHURT CREAM BUTTER GHEE HONEY
JAM VINEGAR YEAST SALT SUGAR FLOUR RICE PASTA NOODLE NOODLES CORN BEAN BEANS
PEA PEAS OLIVE OLIVES NUT NUTS DATE DATES CAKE CAKES COOKIE COOKIES BISCUIT
BISCUITS CANDY CANDIES GUMMIES GUMMY CHOCOLATE POPCORN POPCORNS CEREAL CUSTARD
PIZZA PIZZAS BURGER BURGERS SAUSAGE SAUSAGES NUGGET NUGGETS KEBAB KEBABS
MEATBALL MEATBALLS ICE ICED GRAPE LEAVES LEAVE GRATED HARD SOFT JAR PORTION
SHARING PACKS PACK STILL SPARKLING SPORTS NUTRITION DIABETIC HEALTH FIBER
GATEAUX PIE PIES DESSERT DESSERTS DECORATIONS BATTER BREADCRUMBS CORNED
LUNCHEON BASMATI JASMINE EGYPTIAN VERMICELLI COUSCOUS KARAK ZINGERS LABNEH
GHEE PULSES LENTILS PINEAPPLE MUSHROOM TOMATO TOMATOES PUREE KERNEL HAMMOUS
THAHINA BAKED CAULIFLOWER SPARKLING ETHNIC MEALS READY ROAST SPECIALITY
QUICK INDIV INSTANT PRESERVES CARAMEL FILLED PLANT BASED SUNFLOWER CANOLA
BLENDED OLIVE OIL OILS CARBONATED TAZ TAZWEED NPD SAMPLE SAMPLES BASIC TEST
HYGIENE MASALA SACHET SACHETS FAMILY HOME IN LOOSE BAG BAGS TAKE IMPULSE
OFFAL OFFLAS CARCASS CATERING FPP SBU RETAIL STOCK TRANSFER LIQUIDATION
FOODSERVICE FOOD SERVICE COMBO INT NON-FOOD FOODS DRY GRAINS GRAIN CEREALS
FLAVOUR FLAVOURS FLAVOR FLAVORS EXTRA JUMBO REGULAR SMALL MEDIUM LARGE MINI
CO OP LTD PLC INC GROUP HOLDING TRADING GENERAL EST ESTABLISHMENT
NET BLOCK FONDANT COCOA KIT LIVE FORMED KING QUEEN PACK PACKS PIECE PIECES
PC CB EA UNIT UNITS TRAY TRAYS BOX BOXES CASE CASES CARTON CTN
""".split()

# countries / nationalities / geographic abbreviations - not brands
COUNTRY_TERMS = """
US USA U S UK TH THAI THAILAND AUST AUS AUSTRALIA AUSTRALIAN BRAZIL BRAZILIAN
BRZ INDIA INDIAN IND IRAN IRANIAN OMAN OMANI PAK PAKISTAN PAKISTANI CHINA
CHINESE RUS RUSSIA RUSSIAN UKRAINE UKRAINIAN POL POLAND POLISH PLND ARG
ARGENTINA ARGENTINIAN URUGUAY VIETNAM VIETNAMESE PHILIPPINES PHILIPPINE
PHILLPINO BANGLA BANGLADESH BANGLADESHI EGYPT EGYPTIAN JOR JORDAN JORDANIAN
KSA SAUDI QAR QATAR QATARI KUWAIT KUWAITI KW LEBANON LEBANESE SYRIA SYRIAN
MOROCCO MOROCCAN TUNISIA TUNISIAN GERMANY GERMAN FRANCE FRENCH ITALY ITALIAN
SPAIN SPANISH NETHERLANDS DUTCH HLND BELGIUM BELGIAN DENMARK DANISH NORWAY
NORWEGIAN SWEDEN SWEDISH FINLAND FINNISH NZ NEW ZEALAND CANADA CANADIAN
MEXICO MEXICAN JAPAN JAPANESE KOREA KOREAN GEO GEORGIA GEORGIAN TUR TURKISH
UAE EMIRATI GULF GCC AFG AFGHANISTAN OMN CHI TAZ S. SOUTH NORTH EAST WEST
""".split()

# fish / seafood species - not brands
SEAFOOD_TERMS = """
KINGFISH TILAPIA HALIBUT PIKE OYSTER OYSTERS SALMON MACKEREL SARDINE SARDINES
COD HADDOCK SEABASS BASS GROUPER SNAPPER POMFRET HAMOUR CRAB LOBSTER SQUID
OCTOPUS MUSSEL MUSSELS CLAM CLAMS ANCHOVY HERRING TROUT CATFISH EEL VANNAMI
VANNAMEI GULSHA FISH SHRIMP SHRIMPS PRAWN PRAWNS ROHU HILSA KATLA PANGAS
BOAL RUI NILE PERCH
""".split()

# fruit / common flavour words - high false-positive risk if treated as brand
FLAVOUR_TERMS = """
ORANGE APPLE BANANA MANGO STRAWBERRY BLUEBERRY RASPBERRY BLACKBERRY
WATERMELON MELON PEACH PEAR CHERRY KIWI GUAVA PAPAYA COCONUT VANILLA
CHOCO CHOC PISTACHIO HAZELNUT ALMOND WALNUT CASHEW PEANUT SESAME GINGER
GARLIC ONION PEPPER CHILI CHILLI CUMIN PAPRIKA TURMERIC CINNAMON NUTMEG
CLOVE CARDAMOM LIME
""".split()

# cheese / dairy TYPE names - describe the product, not the brand
CHEESE_TYPE_TERMS = """
MOZZARELLA CHEDDAR GOUDA EDAM PARMESAN RICOTTA FETA HALLOUMI EMMENTAL
GRUYERE BRIE CAMEMBERT KASHKAVAL PROVOLONE MASCARPONE PANEER
""".split()

# generic marketing adjectives / overloaded single words - not brands
MARKETING_TERMS = """
SUPER ROYAL DELUXE SPECIAL GOLD GOLDEN ELITE GRAND MEGA ULTRA PRIME NAT
NATURAL NO1 TOP BEST NEW STAR SIGNATURE HOME BASIC BLUE SWEET SMOKED
""".split()

# Gulf / Middle-East place names commonly used as origin descriptors
PLACE_TERMS = """
DOHA DUBAI ABUDHABI RIYADH JEDDAH MUSCAT MANAMA SHARJAH AJMAN FUJAIRAH
""".split()

CURATED_GENERIC += (COUNTRY_TERMS + SEAFOOD_TERMS + FLAVOUR_TERMS +
                     CHEESE_TYPE_TERMS + MARKETING_TERMS + PLACE_TERMS)


def _category_vocab(df):
    """Every word used inside the 135 lulu_category names is, by
    definition, category language rather than brand language."""
    vocab = set()
    for val in df["lulu_category"].dropna().unique():
        vocab.update(words(val))
    return vocab


def build_generic_terms(df):
    terms = set(CURATED_GENERIC)
    terms.update(_category_vocab(df))
    return terms


# tokens that commonly sit in FRONT of the real brand in material_desc and
# should be skipped when looking at "the starting word(s)" of a description
LEADING_SKIP_TOKENS = set("""
FZ CH FR SF BF KW JOR UAE USA AUS BRZ IND UK NZ KSA IQF RTE RTC WHL BL SL
BNLS SKLS IH CKN CHK BF LMB MTN VL POL ARG PLND TUR CHINA RUS OMN QAR THAI
KOREAN GEO S. S CO CO. FZN
""".split())


def is_all_generic(token_words, generic_terms):
    return len(token_words) > 0 and all(w in generic_terms for w in token_words)


# ==========================================================================
# 3. KNOWN-BRAND DICTIONARY
# ==========================================================================
def build_brand_dictionary(df, crosswalk_df, generic_terms,
                            min_support_non_barcode=MIN_SUPPORT_NON_BARCODE):
    """
    Source of truth = every already-confirmed (non-UNKNOWN) brand_standardized
    value in the master sheet, MINUS anything that is purely generic wording
    (this filters out pre-existing noise like PASTA / DATES / LABNEH /
    'CH AUS WAGYU BEEF' that a previous automated pass mis-tagged as a
    "brand"). As a second safety net, a candidate that was only ever
    produced by a single low-support, non-barcode-verified guess (a one-off
    'llm' tag like 'CH PF' or 'NET') is dropped unless it recurs at least
    `min_support_non_barcode` times.
    brand_crosswalk raw_brand spellings are added as aliases pointing at the
    standardized display name.
    Returns (brand_dict {tight_key: display_name}, sorted word-lengths present).
    """
    brand_dict = {}
    lengths = set()

    conf_df = df[df["brand_standardized"].astype(str).str.upper() != "UNKNOWN"]
    support = conf_df.groupby("brand_standardized").size()
    has_barcode = conf_df.groupby("brand_standardized")["brand_method"].apply(
        lambda s: (s == "barcode").any())

    for b in conf_df["brand_standardized"].dropna().unique():
        b_words = words(b)
        if not b_words or is_all_generic(b_words, generic_terms):
            continue
        if support.get(b, 0) < min_support_non_barcode and not has_barcode.get(b, False):
            continue
        key = tight(b)
        if len(key) < 2:
            continue
        brand_dict[key] = str(b).strip()
        lengths.add(len(b_words))

    if crosswalk_df is not None:
        for _, row in crosswalk_df.iterrows():
            raw, std = row.get("raw_brand"), row.get("standardized_brand")
            if pd.isna(raw) or pd.isna(std):
                continue
            key = tight(raw)
            if len(key) < 2:
                continue
            brand_dict[key] = str(std).strip()
            lengths.add(len(words(raw)))

    return brand_dict, (sorted(lengths, reverse=True) if lengths else [1])


METHOD_WEIGHT = {"barcode": 4, "rule_exact": 3, "rule_fuzzy": 2, "llm": 1}


def build_mgrp_consensus(df, generic_terms):
    """
    Step A1 - the strongest signal: for each mgrp_descr, take a
    reliability-weighted vote of already-confirmed sibling rows'
    brand_standardized values (barcode > rule_exact > rule_fuzzy > llm).
    Returns {mgrp_descr: (brand, evidence_str)} for groups with a clear
    (>= CONSENSUS_SHARE) winner.
    """
    confirmed = df[df["brand_standardized"].astype(str).str.upper() != "UNKNOWN"]
    scores = defaultdict(lambda: defaultdict(float))
    counts = defaultdict(lambda: defaultdict(int))
    for mgrp, brand, method in zip(confirmed["mgrp_descr"], confirmed["brand_standardized"],
                                    confirmed["brand_method"]):
        if not mgrp or is_all_generic(words(brand), generic_terms):
            continue
        w = METHOD_WEIGHT.get(method, 1)
        scores[mgrp][brand] += w
        counts[mgrp][brand] += 1

    consensus = {}
    for mgrp, brand_scores in scores.items():
        best_brand = max(brand_scores, key=brand_scores.get)
        best_score = brand_scores[best_brand]
        total_score = sum(brand_scores.values())
        if best_score / total_score >= CONSENSUS_SHARE:
            n = counts[mgrp][best_brand]
            consensus[mgrp] = (best_brand,
                                f"{n} sibling row(s) sharing mgrp_descr='{mgrp}' already confirmed as '{best_brand}'")
    return consensus


# ==========================================================================
# 4. BRAND RESOLUTION CASCADE
# ==========================================================================
def resolve_from_mgrp_descr(mgrp_descr, brand_dict, generic_terms, max_ngram):
    """Step A2. Try the whole mgrp_descr, then its punctuation-separated
    segments/n-grams (handles 'BK INGR-PURATOS' -> PURATOS)."""
    if not mgrp_descr:
        return None, None

    whole_key = tight(mgrp_descr)
    if whole_key in brand_dict and not is_all_generic(words(mgrp_descr), generic_terms):
        return brand_dict[whole_key], mgrp_descr.strip()

    toks = words(mgrp_descr)
    best, best_len = None, 0
    for _, n, gram in ngrams(toks, max_n=max_ngram):
        if is_all_generic(gram, generic_terms):
            continue
        key = tight(" ".join(gram))
        if key in brand_dict:
            cand_text = " ".join(gram)
            if len(cand_text) > best_len:
                best, best_len = brand_dict[key], len(cand_text)
    if best:
        return best, mgrp_descr.strip()
    return None, None


def resolve_from_desc_prefix(desc_tokens, brand_dict, generic_terms, max_ngram):
    """Step B. Skip leading logistics/state tokens, then check the next
    1..max_ngram words against the dictionary."""
    i, skipped = 0, 0
    while i < len(desc_tokens) and desc_tokens[i] in LEADING_SKIP_TOKENS and skipped < 3:
        i += 1
        skipped += 1

    for start in (0, i):  # try with AND without skipping, longest n-gram first
        sub = desc_tokens[start:start + max_ngram]
        for n in range(len(sub), 0, -1):
            gram = sub[:n]
            if is_all_generic(gram, generic_terms):
                continue
            key = tight(" ".join(gram))
            if key in brand_dict:
                return brand_dict[key], " ".join(gram)
    return None, None


def resolve_from_desc_scan(desc_tokens, brand_dict, generic_terms, max_ngram):
    """Step C. Scan the whole description for any known brand n-gram.
    Longest match wins, and once a span of words is claimed by a match,
    any shorter match nested inside that same span is discarded (so
    'AL NOOR' beats the nested 'NOOR' rather than the two being treated
    as two different, conflicting brands). If more than one distinct,
    non-overlapping brand still shows up, it's a genuine ambiguity."""
    claimed = set()
    found = {}
    for _, n, gram in ngrams(desc_tokens, max_n=max_ngram):
        if any(UNIT_TOKEN_RE.match(w) for w in gram):
            continue
        if is_all_generic(gram, generic_terms):
            continue
        text = " ".join(gram)
        key = tight(text)
        if len(key) < 3 or key not in brand_dict:
            continue
        for start in range(0, len(desc_tokens) - n + 1):
            if desc_tokens[start:start + n] == gram:
                span = set(range(start, start + n))
                if span & claimed:
                    break  # overlaps an already-claimed longer match
                claimed.update(span)
                brand = brand_dict[key]
                if brand not in found or len(text) > len(found[brand]):
                    found[brand] = text
                break

    if len(found) == 1:
        (brand, text), = found.items()
        return brand, text
    if len(found) > 1:
        return "AMBIGUOUS", "; ".join(f"{b}~{t}" for b, t in found.items())
    return None, None


def resolve_brand_row(mgrp_descr, material_desc, brand_dict, generic_terms, max_ngram, mgrp_consensus):
    """Runs the full A1 -> A2 -> B -> C cascade for one row."""
    if mgrp_descr in mgrp_consensus:
        brand, ev = mgrp_consensus[mgrp_descr]
        return brand, "mgrp_descr_sibling_consensus", ev

    brand, ev = resolve_from_mgrp_descr(mgrp_descr, brand_dict, generic_terms, max_ngram)
    if brand:
        return brand, "mgrp_descr_text_match", f"mgrp_descr='{mgrp_descr}' -> matched '{ev}'"

    toks = words(material_desc)
    brand, ev = resolve_from_desc_prefix(toks, brand_dict, generic_terms, max_ngram)
    if brand:
        return brand, "desc_prefix_match", f"leading words of material_desc matched '{ev}'"

    brand, ev = resolve_from_desc_scan(toks, brand_dict, generic_terms, max_ngram)
    if brand == "AMBIGUOUS":
        return "UNKNOWN", "ambiguous_multiple_brands", f"multiple candidate brands found: {ev}"
    if brand:
        return brand, "desc_scan_match", f"found '{ev}' inside material_desc"

    return "UNKNOWN", "unresolved", "no known brand pattern found in mgrp_descr or material_desc"


def run_brand_resolution(df, crosswalk):
    generic_terms = build_generic_terms(df)
    brand_dict, lengths = build_brand_dictionary(df, crosswalk, generic_terms)
    max_ngram = min(max(lengths) if lengths else 3, 4)
    mgrp_consensus = build_mgrp_consensus(df, generic_terms)

    is_unknown = df["brand_standardized"].astype(str).str.upper() == "UNKNOWN"
    final_brand = df["brand_standardized"].astype(str).copy()
    method = pd.Series("existing", index=df.index, dtype=object)
    evidence = pd.Series("", index=df.index, dtype=object)

    for i in df.index[is_unknown]:
        brand, m, ev = resolve_brand_row(
            df.at[i, "mgrp_descr"], df.at[i, "material_desc"],
            brand_dict, generic_terms, max_ngram, mgrp_consensus,
        )
        final_brand.at[i], method.at[i], evidence.at[i] = brand, m, ev

    return final_brand, method, evidence


# ==========================================================================
# 5. CATEGORY AUDIT
# ==========================================================================
CATEGORY_STOPWORDS = set(CURATED_GENERIC) | {"AND", "OR", "THE", "OF", "WITH", "FOR", "IN", "ON", "NO"}

NON_FOOD_MGRP_SIGNALS = [
    "CLEANING", "DETERGENT", "DISPOSABLE", "MCHN", "MACHINE", "BARISTA",
    "ACCESSORIES", "ACCS", "TOOLS", "PACKING", "ALUMINIUM", "FOIL", "HYGIENE",
    "INSECTICIDE", "SOAP", "DIAPER", "BODY CARE", "BABY CARE", "RECHARGE",
    "SELPAK", "CHARCOAL", "NON FOOD", "NON-FOOD", "NONFOOD", "PACKAGING",
    "EQUIPMENT", "GLOVES", "TISSUE", "WIPES", "NAPKIN", "LIGHTER", "MATCHES",
    "BATTERIES", "STATIONERY", "ELECTRONIC", "COFFEE MCHN",
]


def is_non_food_mgrp(mgrp_descr):
    if not isinstance(mgrp_descr, str):
        return False
    s = mgrp_descr.upper()
    return any(sig in s for sig in NON_FOOD_MGRP_SIGNALS)


def build_category_profiles(df, trust_mask):
    """Keyword -> Counter({category: count}) built ONLY from rows we
    already trust. A word seen under many categories counts for less
    (crude IDF) than one that is unique to a single category."""
    cat_word_counts = defaultdict(Counter)
    for _, row in df.loc[trust_mask, ["material_desc", "lulu_category"]].iterrows():
        cat = row["lulu_category"]
        if not cat or cat == "UNKNOWN":
            continue
        for w in set(words(row["material_desc"])):
            if w in CATEGORY_STOPWORDS or UNIT_TOKEN_RE.match(w):
                continue
            cat_word_counts[w][cat] += 1
    word_categories_touched = {w: len(c) for w, c in cat_word_counts.items()}
    return cat_word_counts, word_categories_touched


def suggest_category(material_desc, mgrp_descr, cat_word_counts, word_categories_touched):
    """Specificity-weighted word-overlap score against every candidate
    category. Returns (best_category, best_score, second_score)."""
    text_words = set(words(material_desc)) | set(words(mgrp_descr))
    scores = Counter()
    for w in text_words:
        if w in CATEGORY_STOPWORDS or UNIT_TOKEN_RE.match(w):
            continue
        cats = cat_word_counts.get(w)
        if not cats:
            continue
        weight = 1.0 / word_categories_touched.get(w, 1)
        for cat, cnt in cats.items():
            scores[cat] += weight

    if not scores:
        return None, 0.0, 0.0
    ranked = scores.most_common(2)
    best_cat, best_score = ranked[0]
    second_score = ranked[1][1] if len(ranked) > 1 else 0.0
    return best_cat, best_score, second_score


def run_category_audit(df):
    conf = df["similarity_confidence_score"].fillna(0.0)
    is_barcode = df["category_method"] == "barcode"
    is_unknown_cat = df["lulu_category"].astype(str).str.upper() == "UNKNOWN"
    is_non_food = df["mgrp_descr"].apply(is_non_food_mgrp) | (df["division"] == "Non Foods")

    trust_mask = (is_barcode | (conf >= HIGH_CONF)) & (~is_unknown_cat) & (~is_non_food)
    cat_word_counts, word_cat_touch = build_category_profiles(df, trust_mask)

    final_cat = df["lulu_category"].astype(str).copy()
    flag = pd.Series("", index=df.index, dtype=object)
    evidence = pd.Series("", index=df.index, dtype=object)

    flag.loc[is_barcode] = "TRUSTED_BARCODE"
    flag.loc[(~is_barcode) & (conf >= HIGH_CONF) & (~is_unknown_cat)] = "OK_HIGH_CONFIDENCE"
    flag.loc[(~is_barcode) & (conf >= MED_CONF) & (conf < HIGH_CONF) & (~is_unknown_cat)] = "OK_MEDIUM_CONFIDENCE"
    # The 135-category list is a FOOD taxonomy - a non-food row has no home
    # in it, so UNKNOWN is the *correct* answer, not something to force-fit.
    flag.loc[(flag == "") & is_non_food] = "NOT_FOOD_NO_CATEGORY_APPLICABLE"

    for i in df.index[flag == ""]:
        best_cat, best_score, second_score = suggest_category(
            df.at[i, "material_desc"], df.at[i, "mgrp_descr"], cat_word_counts, word_cat_touch)
        current = df.at[i, "lulu_category"]

        if best_cat is None or best_score < MIN_SCORE:
            flag.at[i] = "NEEDS_MANUAL_REVIEW_NO_MATCH"
            evidence.at[i] = "no category scored high enough among existing categories"
            continue

        clear_winner = (second_score == 0) or (best_score / max(second_score, 1e-9) >= CLEAR_WINNER_RATIO)
        if not clear_winner:
            flag.at[i] = "NEEDS_MANUAL_REVIEW_AMBIGUOUS"
            evidence.at[i] = f"best='{best_cat}'({best_score:.2f}) vs runner-up({second_score:.2f}) too close"
            continue

        if best_cat == current:
            flag.at[i] = "OK_LOW_CONFIDENCE_CONFIRMED"
            evidence.at[i] = f"keyword match agrees with existing category (score {best_score:.2f})"
        else:
            flag.at[i] = "REASSIGNED_NEAREST_MATCH"
            final_cat.at[i] = best_cat
            evidence.at[i] = (f"was '{current}' -> nearest existing category '{best_cat}' "
                               f"(score {best_score:.2f} vs {second_score:.2f})")

    return final_cat, flag, evidence


# ==========================================================================
# 6. WORKBOOK OUTPUT
# ==========================================================================
def autosize_and_style(ws, header_row=1, freeze="A2"):
    header_font = Font(bold=True, name="Arial", color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[header_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells[:200])
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)


def main(in_path, out_path):
    xl = pd.ExcelFile(in_path)
    df = xl.parse("standardized_master")
    crosswalk = xl.parse("brand_crosswalk") if "brand_crosswalk" in xl.sheet_names else None

    print("Rows:", len(df))
    print("Running brand resolution ...")
    final_brand, b_method, b_evidence = run_brand_resolution(df, crosswalk)

    print("Running category audit ...")
    final_cat, c_flag, c_evidence = run_category_audit(df)

    out = df.copy()
    out["brand_final"] = final_brand
    out["brand_resolution_method"] = b_method
    out["brand_resolution_evidence"] = b_evidence
    out["lulu_category_final"] = final_cat
    out["category_audit_flag"] = c_flag
    out["category_audit_evidence"] = c_evidence

    n_total = len(out)
    was_unknown = int((df["brand_standardized"].astype(str).str.upper() == "UNKNOWN").sum())
    now_unknown = int((out["brand_final"].astype(str).str.upper() == "UNKNOWN").sum())
    resolved = was_unknown - now_unknown
    brand_method_counts = b_method.value_counts()
    cat_flag_counts = c_flag.value_counts()

    print(f"Brand: was UNKNOWN for {was_unknown} rows, resolved {resolved}, still unknown {now_unknown}")
    print(brand_method_counts)
    print("\nCategory audit flags:")
    print(cat_flag_counts)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        out.to_excel(writer, sheet_name="standardized_master", index=False)
        if crosswalk is not None:
            crosswalk.to_excel(writer, sheet_name="brand_crosswalk", index=False)

        review_cols = [
            "material_key", "material_code", "material_desc", "mgrp_descr",
            "brand_final", "brand_resolution_method", "brand_resolution_evidence",
            "lulu_category_final", "category_audit_flag", "category_audit_evidence",
        ]
        review = out[
            (out["brand_final"].astype(str).str.upper() == "UNKNOWN") |
            (out["category_audit_flag"].astype(str).str.startswith("NEEDS_MANUAL_REVIEW"))
        ][review_cols]
        review.to_excel(writer, sheet_name="needs_manual_review", index=False)

        notes = [
            "METHODOLOGY (all rule-based - no ML / LLM calls made by this script)",
            "",
            "BRAND: UNKNOWN rows are resolved in priority order -",
            "  1) mgrp_descr sibling consensus - reliability-weighted vote of",
            "     other rows sharing the same material group that already have",
            "     a confirmed brand (barcode > rule_exact > rule_fuzzy > llm).",
            "  2) mgrp_descr text match against every already-confirmed brand.",
            "  3) leading word(s) of material_desc (after skipping FZ/CH/USA/...).",
            "  4) full scan of material_desc for any recognized brand.",
            "A blacklist of generic/category/country/species/marketing words",
            "(auto-derived from your 135 category names + curated additions)",
            "stops those being mistaken for brands. A row is only left UNKNOWN",
            "when none of the 4 steps find anything - see the",
            "'needs_manual_review' sheet and the brand_resolution_* columns.",
            "",
            "CATEGORY: barcode-matched and high-confidence rows are untouched.",
            "Rows whose material group is clearly non-food (cleaning supplies,",
            "machine parts, disposables...) are marked not-applicable, since",
            "the 135-category list is food-only and UNKNOWN is correct there.",
            "Everything else is scored against keyword profiles built only from",
            "trusted rows, and reassigned to an EXISTING category only when",
            "there is a clear, unambiguous winner - never a new category.",
            "",
            "Every filled/flagged row carries a *_method / *_evidence column",
            "so any assignment can be checked and overridden by a human in",
            "seconds. Tune thresholds/word lists at the top of the script.",
            "",
        ]
        summary_rows = [("NOTE", n) for n in notes]
        summary_rows += [
            ("Total rows", n_total), ("", ""),
            ("BRAND RESOLUTION", ""),
            ("Brand UNKNOWN before", was_unknown),
            ("Brand UNKNOWN after", now_unknown),
            ("Newly resolved", resolved),
        ]
        summary_rows += [(f"  method: {k}", int(v)) for k, v in brand_method_counts.items()]
        summary_rows += [("", ""), ("CATEGORY AUDIT", "")]
        summary_rows += [(f"  flag: {k}", int(v)) for k, v in cat_flag_counts.items()]
        pd.DataFrame(summary_rows, columns=["Metric", "Value"]).to_excel(
            writer, sheet_name="summary", index=False)

    wb = load_workbook(out_path)
    for name in ["standardized_master", "brand_crosswalk", "needs_manual_review", "summary"]:
        if name in wb.sheetnames:
            autosize_and_style(wb[name])

    ws = wb["standardized_master"]
    headers = [c.value for c in ws[1]]
    col_brand = headers.index("brand_final") + 1
    col_flag = headers.index("category_audit_flag") + 1
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    amber_fill = PatternFill("solid", fgColor="FFEB9C")
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, col_brand).value == "UNKNOWN":
            ws.cell(r, col_brand).fill = red_fill
        flag_val = ws.cell(r, col_flag).value
        if isinstance(flag_val, str) and flag_val.startswith("NEEDS_MANUAL_REVIEW"):
            ws.cell(r, col_flag).fill = amber_fill
        elif flag_val == "REASSIGNED_NEAREST_MATCH":
            ws.cell(r, col_flag).fill = green_fill

    wb.save(out_path)
    print("Saved:", out_path)


if __name__ == "__main__":
    in_path = sys.argv[1] if len(sys.argv) > 1 else "dim_material_master_qnie_csv.xlsx"
    out_path = sys.argv[2] if len(sys.argv) > 2 else "dim_material_master_qnie_csv_RESOLVED.xlsx"
    main(in_path, out_path)